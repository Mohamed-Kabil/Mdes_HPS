#!/usr/bin/env python3
"""
mdes_cs_prereleases.py — checks Mastercard's MDES Customer Service
pre-release notes for mentions of the 6 tracked CS operations (Search,
Token Activate, Update, Suspend, Unsuspend, Delete).

Same relationship to mdes_cs_divergence_report.py as
phase2_job_a_new_release_alert.py has to phase1_historical_audit.py: this
reports what a Mastercard-announced upcoming release *says* will change,
not whether the Java implementation already reflects it (that's still
mdes_cs_divergence_report.py's job, against the *current live* spec).

The CS release-history index page uses a different table shape than the
Pre-Digitization one (2 columns: title+link, date — no separate "MDES
release"/"note type" columns), so this doesn't reuse
fetch_mc_source.parse_release_notes_table() as-is; it has its own parser
below, same '/index.md' trick (any developer.mastercard.com page returns
clean Markdown at that suffix, no HTML scraping needed).

Deliberately simple keyword matching (same philosophy as
phase2_job_a_new_release_alert.py's find_matching_apis: "intentionally
basic string matching, not NLP — good enough to prove the fetch + filter
works", not a claim of perfect recall).

Usage:
    python3 mdes_cs_prereleases.py                # latest release note only
    python3 mdes_cs_prereleases.py --refresh
    python3 mdes_cs_prereleases.py --limit 5       # check the 5 most recent
"""

import argparse
import json
import os
import re
import sys
import urllib.request
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(HERE, 'mc_divergence'))

import fetch_mc_source  # reused only for USER_AGENT + the '/index.md' fetch pattern
from parse_release_note import parse_note
# Checkpoint machinery + date parsing are generic (cache_dir-parameterized,
# no Pre-Digitization-specific state) -- reused as-is rather than
# reimplemented, same rationale as reusing phase1's openpyxl styling
# helpers in mdes_cs_divergence_report.py.
from phase2_job_a_new_release_alert import (
    load_checkpoint, save_checkpoint, advance_checkpoint, releases_since_checkpoint,
    parse_release_date, HEADER_FILL, HEADER_FONT, TITLE_FONT, WRAP,
    _safe_sheet_name, _style_header_row, _autosize,
)
import send_email as send_email_module

RELEASE_NOTES_INDEX_URL = "https://developer.mastercard.com/mdes-customer-service/documentation/release-history/index.md"

DEFAULT_CACHE_DIR = os.path.join(HERE, 'mc_divergence', 'cache', 'cs')

TARGET_APIS = ["Search", "Token Activate", "Token Update", "Token Suspend", "Token Unsuspend", "Token Delete"]

API_KEYWORDS = {
    "Search": ["Search API", "Token Search"],
    "Token Activate": ["Token Activate", "Token Activation"],
    "Token Update": ["Token Update"],
    "Token Suspend": ["Token Suspend"],
    "Token Unsuspend": ["Token Unsuspend"],
    "Token Delete": ["Token Delete"],
}

# | [Title](url) | Date |
TABLE_ROW_RE = re.compile(
    r'^\|\s*\[(?P<title>[^\]]+)\]\((?P<url>[^)]+)\)\s*\|\s*(?P<upgrade_date>[^|]+?)\s*\|\s*$'
)


def _fetch(url):
    req = urllib.request.Request(url, headers={"User-Agent": fetch_mc_source.USER_AGENT})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read().decode("utf-8")


def parse_release_notes_table(markdown_text):
    entries = []
    for line in markdown_text.splitlines():
        m = TABLE_ROW_RE.match(line.strip())
        if not m:
            continue
        entries.append({
            "title": m.group("title").strip(),
            "url": m.group("url").strip(),
            "upgrade_date": m.group("upgrade_date").strip(),
        })
    return entries


def fetch_release_notes(cache_dir):
    text = _fetch(RELEASE_NOTES_INDEX_URL)
    entries = parse_release_notes_table(text)

    out_path = os.path.join(cache_dir, "known_releases.json")
    os.makedirs(cache_dir, exist_ok=True)
    payload = {
        "source_url": RELEASE_NOTES_INDEX_URL,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "count": len(entries),
        "releases": entries,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    print(f"  [fetch-cs] release notes table -> {out_path} ({len(entries)} entries)", file=sys.stderr)
    if not entries:
        print("  [warn] 0 entries parsed — table format may have changed, "
              "check the raw markdown before trusting downstream results", file=sys.stderr)
    return entries


def get_release_notes_table(cache_dir, refresh=False):
    cache_path = os.path.join(cache_dir, "known_releases.json")
    if refresh or not os.path.exists(cache_path):
        return fetch_release_notes(cache_dir)
    with open(cache_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data["releases"]


def _slug_from_url(url):
    return url.rstrip("/").split("/")[-2] if url.rstrip("/").endswith("index.md") else \
        url.rstrip("/").split("/")[-1]


def fetch_note_text(entry, cache_dir):
    raw_dir = os.path.join(cache_dir, "notes_raw")
    os.makedirs(raw_dir, exist_ok=True)
    path = os.path.join(raw_dir, f"{_slug_from_url(entry['url'])}.md")

    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return f.read()

    text = _fetch(entry["url"])
    with open(path, "w", encoding="utf-8") as f:
        f.write(text)
    return text


def find_matching_apis(note_text):
    matched = []
    for api in TARGET_APIS:
        keywords = API_KEYWORDS[api]
        if any(re.search(re.escape(kw), note_text, re.IGNORECASE) for kw in keywords):
            matched.append(api)
    return matched


def check_latest(cache_dir, refresh=False):
    """Returns the newest release-note entry (table is published
    newest-first) augmented with matched_apis, or None if the table is
    empty. Read-only preview -- no checkpoint involved, same role as
    /api/phase2/latest for Part 1."""
    entries = get_release_notes_table(cache_dir, refresh=refresh)
    if not entries:
        return None
    entry = entries[0]
    text = fetch_note_text(entry, cache_dir)
    matched_apis = find_matching_apis(text)
    return {**entry, "matched_apis": matched_apis}


# ============================================================================
# Checkpoint-based two-check pattern — same relationship as
# phase2_job_a_new_release_alert.py's check_pending_since_predig() /
# check_since_checkpoint(): two independently-triggerable checks sharing one
# persisted checkpoint (mc_divergence/cache/cs/phase2_checkpoint.json,
# isolated from Part 1's own checkpoint by living under a different
# cache_dir), not automatically switching between each other.
#
#   - check_pending() ("Check after last CS audit"): every release
#     whose Production date (parsed from the note body itself, via
#     parse_release_note.parse_note -- reused as-is, its date-line regex
#     already matches this note format) hasn't happened yet -- i.e. not yet
#     reflected in the live mdes-customer-service.yaml spec, same
#     "Production date <= today implies already live" assumption Part 1
#     uses. The ONLY action that ADVANCES the checkpoint.
#   - check_since_checkpoint() ("Check latest CS release"): releases
#     newer than whatever check_pending() last examined. Never advances the
#     checkpoint itself. Falls back to just the newest table entry if
#     check_pending() has never been run.
#
# Unlike Part 1, there's no cheap month/year pre-filter before fetching each
# note (the CS table has no separate "MDES release" column to filter on
# cheaply) -- the table only has ~15 entries total, so fetching/parsing all
# of them is negligible.
# ============================================================================

def releases_pending_production(entries, cache_dir, today=None):
    if today is None:
        today = datetime.now().date()
    pending = []
    for entry in entries:
        text = fetch_note_text(entry, cache_dir)
        parsed = parse_note(text, url=entry["url"])
        prod_date_raw = parsed["timeline"].get("production_date")
        prod_date = parse_release_date(prod_date_raw)
        if prod_date is None or prod_date > today:
            pending.append({**entry, "production_date": prod_date_raw})
    return pending


def filter_relevant_releases(entries, cache_dir):
    relevant = []
    for entry in entries:
        print(f"  [cs-prereleases] checking {entry['title']} ({entry['url']})", file=sys.stderr)
        text = fetch_note_text(entry, cache_dir)
        matched_apis = find_matching_apis(text)
        if matched_apis:
            parsed = parse_note(text, url=entry["url"])
            relevant.append({**entry, "matched_apis": matched_apis, "parsed": parsed})
    return relevant


def check_pending(entries, cache_dir):
    """Returns {'working_set', 'relevant', 'outcome'}. Caller must call
    advance_checkpoint(cache_dir, entries) afterward -- this function itself
    has no side effects, same contract as Part 1's equivalent."""
    working_set = releases_pending_production(entries, cache_dir)
    relevant = filter_relevant_releases(working_set, cache_dir)

    if not working_set:
        outcome = "no_new"
    elif not relevant:
        outcome = "new_no_impact"
    else:
        outcome = "new_with_impact"

    return {"working_set": working_set, "relevant": relevant, "outcome": outcome}


def check_since_checkpoint(entries, cache_dir):
    """Returns {'working_set', 'relevant', 'outcome', 'used_fallback'}.
    Never advances the checkpoint."""
    checkpoint = load_checkpoint(cache_dir)
    since_checkpoint = releases_since_checkpoint(entries, checkpoint)
    used_fallback = since_checkpoint is None
    working_set = entries[:1] if used_fallback else since_checkpoint

    relevant = filter_relevant_releases(working_set, cache_dir)

    if not working_set:
        outcome = "no_new"
    elif not relevant:
        outcome = "new_no_impact"
    else:
        outcome = "new_with_impact"

    return {"working_set": working_set, "relevant": relevant, "outcome": outcome, "used_fallback": used_fallback}


# ============================================================================
# Email + Excel report — same relationship to check_pending()/
# check_since_checkpoint() as phase2_job_a_new_release_alert.py's own
# email_subject/render_email_body/render_report_xlsx have to its checks.
# ============================================================================

def email_subject(relevant):
    return f"[MDES Customer Service] {len(relevant)} new release(s) impacting tracked operations"


def format_change_block(c, indent="    "):
    lines = [f"{indent}- {c['title']}"]
    if c.get("description"):
        lines.append(f"{indent}  Description : {c['description']}")
    endpoints = c["endpoints"]
    inferred_note = (" (inferred from Impacted APIs, no explicit API Reference found)"
                      if c.get("endpoints_inferred") else "")
    lines.append(f"{indent}  Endpoints   : {', '.join(endpoints) if endpoints else '(none found)'}{inferred_note}")
    if c["fields"]:
        lines.append(f"{indent}  Fields      :")
        for f in c["fields"]:
            parts = [f"name={f['name']}"]
            if f.get("type"):
                parts.append(f"type={f['type']}")
            if f.get("required_raw"):
                parts.append(f"required={f['required_raw']}")
            lines.append(f"{indent}    - " + ", ".join(parts))
    else:
        lines.append(f"{indent}  Fields      : (no structured field table found for this change)")
    return lines


def format_release_block(entry):
    parsed = entry["parsed"]
    timeline = parsed["timeline"]
    lines = [
        f"- {entry['title']}",
        f"    Updated on          : {entry['upgrade_date']}",
        f"    Affects             : {', '.join(entry['matched_apis'])}",
        f"    URL                 : {entry['url']}",
        f"    MTF date            : {timeline['mtf_date']}",
        f"    Production date     : {timeline['production_date']}",
        f"    Changes ({len(parsed['changes'])}) :",
    ]
    if not parsed["changes"]:
        lines.append("      (no changes parsed — see parse_release_note.py's known limitations)")
    for c in parsed["changes"]:
        lines += format_change_block(c)
    lines.append("")
    return lines


def render_email_body(relevant):
    lines = [
        f"{len(relevant)} MDES Customer Service pre-release note(s) mention the tracked operations "
        f"({', '.join(TARGET_APIS)}):",
        "",
    ]
    for entry in relevant:
        lines += format_release_block(entry)
    return "\n".join(lines)


def default_report_xlsx_path():
    return os.path.join(HERE, "mc_divergence", "reports",
                         f"mdes_cs_prerelease_alert - {datetime.now().strftime('%d-%m-%y %Hh%Mm%S')}.xlsx")


def _write_summary_sheet(wb, relevant):
    ws = wb.create_sheet("Summary", 0)
    ws['A1'] = f"MDES Customer Service — {len(relevant)} release(s) affecting tracked operations"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:D1')

    header_row = 3
    headers = ['Note', 'Updated on', 'Affected operations', 'URL']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(headers))

    r = header_row + 1
    for entry in relevant:
        values = [entry['title'], entry['upgrade_date'], ', '.join(entry['matched_apis']), entry['url']]
        for col, v in enumerate(values, start=1):
            ws.cell(row=r, column=col, value=v).alignment = WRAP
        r += 1

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize(ws, [45, 18, 34, 55])


def _write_operation_sheet(wb, sheet_name, api, relevant):
    ws = wb.create_sheet(sheet_name)
    ws['A1'] = api
    ws['A1'].font = TITLE_FONT

    headers = ['Note', 'Updated on', 'Endpoint(s)', 'Change', 'Description',
               'Field', 'Type', 'Required']
    header_row = 3
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(headers))

    row = header_row + 1
    for entry in relevant:
        if api not in entry['matched_apis']:
            continue
        changes = entry['parsed']['changes']
        if not changes:
            values = [entry['title'], entry['upgrade_date'], '', '(no structured change parsed)',
                       '', '', '', '']
            for col, v in enumerate(values, start=1):
                ws.cell(row=row, column=col, value=v).alignment = WRAP
            row += 1
            continue
        for c in changes:
            endpoints = ', '.join(c['endpoints']) if c['endpoints'] else '(no endpoint identified)'
            if not c['fields']:
                values = [entry['title'], entry['upgrade_date'], endpoints, c['title'],
                           c.get('description') or '', '', '', '']
                for col, v in enumerate(values, start=1):
                    ws.cell(row=row, column=col, value=v).alignment = WRAP
                row += 1
                continue
            for f in c['fields']:
                values = [entry['title'], entry['upgrade_date'], endpoints, c['title'],
                           c.get('description') or '', f['name'], f.get('type') or '', f.get('required_raw') or '']
                for col, v in enumerate(values, start=1):
                    ws.cell(row=row, column=col, value=v).alignment = WRAP
                row += 1

    last_row = row - 1
    if last_row >= header_row + 1:
        table_ref = f"A{header_row}:{get_column_letter(len(headers))}{last_row}"
        table = Table(displayName=f"T{re.sub(r'[^A-Za-z0-9]', '', sheet_name)}", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
        ws.add_table(table)

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize(ws, [40, 18, 26, 34, 40, 30, 12, 14])


def render_report_xlsx(relevant, xlsx_path):
    wb = Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb, relevant)
    used_names = set()
    for api in TARGET_APIS:
        if not any(api in entry['matched_apis'] for entry in relevant):
            continue
        sheet_name = _safe_sheet_name(api, used_names)
        _write_operation_sheet(wb, sheet_name, api, relevant)

    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('--cache-dir', default=DEFAULT_CACHE_DIR)
    parser.add_argument('--refresh', action='store_true')
    parser.add_argument('--limit', type=int, default=1, help="check the N most recent release notes (default 1)")
    args = parser.parse_args()

    entries = get_release_notes_table(args.cache_dir, refresh=args.refresh)[:args.limit]
    for entry in entries:
        text = fetch_note_text(entry, args.cache_dir)
        matched = find_matching_apis(text)
        flag = f" -- MATCHES: {', '.join(matched)}" if matched else ""
        print(f"{entry['upgrade_date']}  {entry['title']}{flag}")


if __name__ == '__main__':
    main()
