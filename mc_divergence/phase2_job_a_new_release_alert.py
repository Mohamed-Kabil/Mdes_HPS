#!/usr/bin/env python3
"""
phase2_job_a_new_release_alert.py — Phase 2, job A of the Mastercard MDES
divergence pipeline described in mastercard_divergence_prompt.md.

Pipeline (this script only):
  1. Fetch the current MDES Pre-Digitization pre-release-notes index
     (reuses fetch_mc_source.fetch_release_notes(), same '/index.md' trick —
     see that module for why it returns clean Markdown with no HTML
     scraping needed).
  2. For each row (API version / MDES release / type of note / date of
     upgrading / link to its own .md page), fetch that individual note.
  3. Check the note's text for simple keyword mentions of the 5 tracked
     APIs (TARGET_APIS below) — plain substring matching, no LLM.
  4. For every matching note, also run it through parse_release_note.py's
     full structured parser (same one phase1_historical_audit.py uses) to
     pull out every change's title, description, endpoint(s) and field
     table — not just "this note mentions API X".
  5. Print every matching release (with its full change/field detail) to
     the terminal (or a clear "nothing found" message), write a local
     Markdown email draft under reports/ either way, and — only with
     --send — actually email it via send_email.py.

Deliberately NOT done here:
  - no field-level diff against data.yaml/pre-dig.yaml (diff_openapi_all.py)
    — this reports what the note says changed, not whether data.yaml
    already reflects it
  - no LLM-assisted note parsing

Two independently-triggerable checks (--check pending|latest, default
pending — see section 3quater below), sharing one persisted checkpoint
(cache/phase2_checkpoint.json) but not automatically switching between
each other:
  - --check pending ("Vérifier après le dernier pre-dig"): every release
    whose Production date hasn't happened yet — i.e. not yet reflected in
    pre-dig.yaml, which only ever contains what's already live. Recomputed
    fresh from Production dates every run. The ONLY action that ADVANCES
    the checkpoint — the authoritative sweep that moves the reference
    point forward.
  - --check latest ("Vérifier la dernière release"): releases newer than
    whatever 'pending' last examined. Does NOT advance the checkpoint
    itself, so running it repeatedly keeps re-showing the same still-new
    release(s) until 'pending' runs again. Falls back to just the single
    newest table entry if 'pending' has never been run yet.
Use --all/--limit to widen the check for testing — those modes bypass both
checks entirely and are read-only with respect to the checkpoint.

Usage:
    python3 phase2_job_a_new_release_alert.py                  # --check pending (default)
    python3 phase2_job_a_new_release_alert.py --check latest
    python3 phase2_job_a_new_release_alert.py --send             # ...and actually email it
    python3 phase2_job_a_new_release_alert.py --refresh
    python3 phase2_job_a_new_release_alert.py --limit 5     # latest 5, for testing (checkpoint untouched)
    python3 phase2_job_a_new_release_alert.py --all         # every release, for testing (checkpoint untouched)
    python3 phase2_job_a_new_release_alert.py --cache-dir cache
"""

import argparse
import json
import os
import re
import sys
import urllib.request
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import fetch_mc_source
from parse_release_note import parse_note
import send_email as send_email_module

# The 5 APIs this job watches for (prompt: "APIs prioritaires"), spelled as
# given by the workstream owner. Kept as display-name strings, not paths —
# matching happens against release-note prose, which names APIs by their
# human title ("Authorize Service (AS)"), not their endpoint path.
TARGET_APIS = [
    "RequestActivationMethods",
    "DeliverActivationCode",
    "AuthoriseService",
    "NotifyServiceActivated",
    "NotifyTokenUpdated",
    "ValidateActivationCode",
    "GetAccountInformation",
    "NotifySuspiciousEvents",
]

# Simple keyword matching per API: the camel-case name split into words
# (Mastercard's own prose form, e.g. "Request Activation Methods"), plus the
# parenthesized abbreviation Mastercard notes consistently use for it
# (e.g. "(RAM)") and a spelling variant where relevant (Mastercard's site
# uses the American "Authorize", the workstream's own list uses "Authorise").
# This is intentionally basic string matching, not NLP — good enough to
# "prove the fetch + filter works" per the brief; a real parse (per-change
# API attribution) is parse_release_note.py's job, not this one.
API_KEYWORDS = {
    "RequestActivationMethods": ["Request Activation Method", "(RAM)"],
    "DeliverActivationCode": ["Deliver Activation Code", "(DAC)"],
    "AuthoriseService": ["Authorize Service", "Authorise Service", "(AS)"],
    "NotifyServiceActivated": ["Notify Service Activated", "(NSA)"],
    "NotifyTokenUpdated": ["Notify Token Updated", "(NTU)"],
    "ValidateActivationCode": ["Validate Activation Code", "(VAC)"],
    "GetAccountInformation": ["Get Account Information", "(GAI)"],
    "NotifySuspiciousEvents": ["Notify Suspicious Events", "(NSE)"],
}


# ============================================================================
# 1. Fetch + parse the release-notes index table
# ============================================================================

def get_release_notes_table(cache_dir, refresh=False):
    """Returns the list of {title, url, mdes_release, note_type,
    upgrade_date} dicts from the pre-release-notes index. Reuses
    fetch_mc_source's cache (cache/known_releases.json) so re-running this
    script doesn't re-hit the network unless --refresh is passed."""
    cache_path = os.path.join(cache_dir, "known_releases.json")
    if refresh or not os.path.exists(cache_path):
        fetch_mc_source.fetch_release_notes(cache_dir)
    with open(cache_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data["releases"]


# ============================================================================
# 2. Fetch each individual release note's own .md page
# ============================================================================

def _slug_from_url(url):
    return url.rstrip("/").split("/")[-2] if url.rstrip("/").endswith("index.md") else \
        url.rstrip("/").split("/")[-1]


def fetch_note_text(entry, cache_dir):
    """Fetches (or reuses a cached copy of) one release note's Markdown.
    Cached under cache/notes_raw/<slug>.md, same layout phase1 uses, so the
    two jobs share the cache instead of each re-downloading the same notes."""
    raw_dir = os.path.join(cache_dir, "notes_raw")
    os.makedirs(raw_dir, exist_ok=True)
    path = os.path.join(raw_dir, f"{_slug_from_url(entry['url'])}.md")

    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return f.read()

    req = urllib.request.Request(entry["url"], headers={"User-Agent": fetch_mc_source.USER_AGENT})
    with urllib.request.urlopen(req, timeout=30) as resp:
        text = resp.read().decode("utf-8")
    with open(path, "w", encoding="utf-8") as f:
        f.write(text)
    return text


# ============================================================================
# 3. Keyword filter against the 5 tracked APIs
# ============================================================================

def find_matching_apis(note_text):
    """Plain case-insensitive substring match of each TARGET_APIS entry's
    keyword list against the note's full text. Returns the subset of
    TARGET_APIS that appear to be mentioned, in TARGET_APIS order."""
    matched = []
    for api in TARGET_APIS:
        keywords = API_KEYWORDS[api]
        if any(re.search(re.escape(kw), note_text, re.IGNORECASE) for kw in keywords):
            matched.append(api)
    return matched


def filter_relevant_releases(entries, cache_dir, limit=None):
    """For each release-notes-table entry, fetches its note, checks it
    against the 5 tracked APIs, and — only for matches — also runs the full
    structured parse (parse_release_note.parse_note). Returns only the
    entries with at least one keyword match, each augmented with a
    'matched_apis' list and a 'parsed' record (version/impacted_apis/
    timeline/changes, same shape phase1_historical_audit.py consumes)."""
    if limit is not None:
        entries = entries[:limit]

    relevant = []
    for entry in entries:
        print(f"  [phase2a] checking {entry['title']} ({entry['url']})", file=sys.stderr)
        text = fetch_note_text(entry, cache_dir)
        matched_apis = find_matching_apis(text)
        if matched_apis:
            parsed = parse_note(text, url=entry["url"])
            relevant.append({**entry, "matched_apis": matched_apis, "parsed": parsed})
    return relevant


# ============================================================================
# 3bis. Check 1 — releases not yet reflected in pre-dig.yaml (Production
#       date still in the future). pre-dig.yaml itself is never touched
#       here: same "Production date <= today implies already live in the
#       always-freshly-fetched pre-dig.yaml" assumption
#       phase1_historical_audit.py's own docstring documents. Powers the
#       standalone "check everything after the current pre-dig" action
#       (see check_pending_since_predig() below).
# ============================================================================

# Mirrors phase1_historical_audit.py's parse_mdes_release()/MONTH_NUM —
# duplicated (not imported) so this script stays independently runnable
# without pulling in phase1's heavier diff_openapi_all dependency chain,
# same "these scripts don't touch the network directly" independence the
# module docstring already promises for the rest of this file.
MONTH_NUM = {
    'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
    'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7,
    'aug': 8, 'august': 8, 'sep': 9, 'sept': 9, 'september': 9,
    'oct': 10, 'october': 10, 'nov': 11, 'november': 11, 'dec': 12, 'december': 12,
}
MDES_RELEASE_RE = re.compile(r'^(\d{4})\s+([A-Za-z]+)')


def parse_mdes_release(mdes_release):
    m = MDES_RELEASE_RE.match(mdes_release or '')
    if not m:
        return None
    year = int(m.group(1))
    month = MONTH_NUM.get(m.group(2).strip().lower())
    if month is None:
        return None
    return (year, month)


# Release notes format Production/MTF dates inconsistently ('16 July 2025'
# vs '5 Dec 2020') — try full month name first, then abbreviated.
RELEASE_DATE_FORMATS = ("%d %B %Y", "%d %b %Y")


def parse_release_date(raw):
    """'16 July 2025' / '5 Dec 2020' -> date(...). None if unparseable
    (some older notes have no clean Production date line at all — see
    parse_release_note.py's known limitations)."""
    if not raw:
        return None
    raw = raw.strip()
    for fmt in RELEASE_DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def releases_pending_production(entries, cache_dir, today=None):
    """Check 1: releases whose Production date hasn't happened yet — i.e.
    not yet reflected in pre-dig.yaml. A cheap month/year pre-filter on
    'mdes_release' first skips fetching notes that are certainly already
    live (their own release month has fully elapsed), then the reduced set
    is fetched+parsed for its exact Production date. Unparseable dates are
    treated as still-pending (fail safe — better a false positive than a
    silently dropped upcoming change)."""
    if today is None:
        today = datetime.now().date()
    pending = []
    for entry in entries:
        ym = parse_mdes_release(entry.get("mdes_release"))
        if ym is not None and ym < (today.year, today.month):
            continue
        text = fetch_note_text(entry, cache_dir)
        parsed = parse_note(text, url=entry["url"])
        prod_date_raw = parsed["timeline"].get("production_date")
        prod_date = parse_release_date(prod_date_raw)
        if prod_date is None or prod_date > today:
            pending.append({**entry, "production_date": prod_date_raw})
    return pending


# ============================================================================
# 3ter. Check 2 — checkpoint: "the latest release checked" persisted across
#       runs, so a normal run only looks at what's genuinely NEW since last
#       time instead of re-examining (and re-alerting on) releases already
#       covered by a previous run, even if they're still pending production.
# ============================================================================

def _checkpoint_path(cache_dir):
    return os.path.join(cache_dir, "phase2_checkpoint.json")


def load_checkpoint(cache_dir):
    path = _checkpoint_path(cache_dir)
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_checkpoint(cache_dir, entry):
    with open(_checkpoint_path(cache_dir), "w", encoding="utf-8") as f:
        json.dump({
            "last_checked_url": entry["url"],
            "last_checked_title": entry["title"],
            "checked_at": datetime.now(timezone.utc).isoformat(),
        }, f, indent=2)


def advance_checkpoint(cache_dir, entries):
    """Called once at the end of a normal (non --all/--limit) run — advances
    the checkpoint to the newest release currently in the index, regardless
    of outcome, so nothing examined this run is ever re-examined next time."""
    if entries:
        save_checkpoint(cache_dir, entries[0])


def releases_since_checkpoint(entries, checkpoint):
    """entries is newest-first. Returns every entry strictly newer than
    (i.e. appearing before, in table order) the checkpoint's release. If
    there's no checkpoint, or its URL isn't found in the current table
    (stale/reset state), returns None — the caller falls back to
    releases_pending_production() instead (bootstrap)."""
    if checkpoint is None:
        return None
    for i, entry in enumerate(entries):
        if entry["url"] == checkpoint.get("last_checked_url"):
            return entries[:i]
    return None


# ============================================================================
# 3quater. Two independently-triggerable checks (each its own dashboard
#          button / CLI mode), sharing the checkpoint but not automatically
#          switching between each other:
#
#          - check_pending_since_predig() ("Vérifier après le dernier
#            pre-dig"): check 1 standalone — every release still not
#            reflected in pre-dig.yaml (Production date in the future),
#            recomputed fresh every time from Production dates, independent
#            of the checkpoint. This is the ONLY action that ADVANCES the
#            checkpoint (the caller does it via advance_checkpoint()
#            right after) — it's the authoritative sweep that moves the
#            reference point forward.
#
#          - check_since_checkpoint() ("Vérifier la dernière release"):
#            check 2 standalone — releases newer than whatever the pre-dig
#            check last examined. Deliberately does NOT advance the
#            checkpoint itself, so clicking it repeatedly keeps re-showing
#            the same still-new release(s) until the pre-dig check runs
#            again and moves the checkpoint forward. Falls back to just the
#            single newest table entry if no checkpoint exists yet (pre-dig
#            check has never been run).
# ============================================================================

def check_pending_since_predig(entries, cache_dir):
    """Returns {'working_set': [...], 'relevant': [...],
    'outcome': 'no_new'|'new_no_impact'|'new_with_impact'}. Caller must call
    advance_checkpoint(cache_dir, entries) afterward — this function itself
    has no side effects."""
    working_set = releases_pending_production(entries, cache_dir)
    relevant = filter_relevant_releases(working_set, cache_dir, limit=None)

    if not working_set:
        outcome = "no_new"
    elif not relevant:
        outcome = "new_no_impact"
    else:
        outcome = "new_with_impact"

    return {"working_set": working_set, "relevant": relevant, "outcome": outcome}


def check_since_checkpoint(entries, cache_dir):
    """Returns {'working_set': [...], 'relevant': [...],
    'outcome': 'no_new'|'new_no_impact'|'new_with_impact',
    'used_fallback': bool}. 'used_fallback' is True when there was no
    checkpoint to compare against (check_pending_since_predig() has never
    been run) and this fell back to just the single newest table entry.
    Never advances the checkpoint."""
    checkpoint = load_checkpoint(cache_dir)
    since_checkpoint = releases_since_checkpoint(entries, checkpoint)
    used_fallback = since_checkpoint is None
    working_set = entries[:1] if used_fallback else since_checkpoint

    relevant = filter_relevant_releases(working_set, cache_dir, limit=None)

    if not working_set:
        outcome = "no_new"
    elif not relevant:
        outcome = "new_no_impact"
    else:
        outcome = "new_with_impact"

    return {"working_set": working_set, "relevant": relevant, "outcome": outcome, "used_fallback": used_fallback}


# ============================================================================
# 4. Shared formatting (terminal print AND email body both use this, so the
#    two never drift apart)
# ============================================================================

def format_change_block(c, indent="    "):
    lines = [f"{indent}- {c['title']}"]
    if c.get("description"):
        lines.append(f"{indent}  Description : {c['description']}")
    endpoints = c["endpoints"]
    inferred_note = (" (inferred from Impacted APIs, no explicit API Reference found)"
                      if c.get("endpoints_inferred") else "")
    lines.append(f"{indent}  Endpoints   : {', '.join(endpoints) if endpoints else '(none found)'}{inferred_note}")
    if c["fields"]:
        lines.append(f"{indent}  Fields      :")
        for f in c["fields"]:
            parts = [f"name={f['name']}"]
            if f.get("type"):
                parts.append(f"type={f['type']}")
            if f.get("min"):
                parts.append(f"min={f['min']}")
            if f.get("max"):
                parts.append(f"max={f['max']}")
            if f.get("required_raw"):
                parts.append(f"required={f['required_raw']}")
            if f.get("description"):
                parts.append(f"description={f['description']}")
            lines.append(f"{indent}    - " + ", ".join(parts))
    else:
        lines.append(f"{indent}  Fields      : (no structured field table found for this change)")
    return lines


def format_release_block(entry):
    parsed = entry["parsed"]
    timeline = parsed["timeline"]
    lines = [
        f"- {entry['title']}",
        f"    API version     : {entry['mdes_release']}",
        f"    Type of note    : {entry['note_type']}",
        f"    Date of upgrade : {entry['upgrade_date']}",
        f"    Affects         : {', '.join(entry['matched_apis'])}",
        f"    URL             : {entry['url']}",
        f"    Impacted APIs (as listed in note) : {', '.join(parsed['impacted_apis']) or '(none listed)'}",
        f"    MTF date        : {timeline['mtf_date']}",
        f"    Production date : {timeline['production_date']}",
        f"    Changes ({len(parsed['changes'])}) :",
    ]
    if not parsed["changes"]:
        lines.append(f"{'    '}  (no changes parsed — see parse_release_note.py's known limitations)")
    for c in parsed["changes"]:
        lines += format_change_block(c)
    lines.append("")
    return lines


# ============================================================================
# 5. Terminal output
# ============================================================================

def print_relevant_releases(relevant):
    if not relevant:
        print("No relevant releases found — none of the tracked APIs "
              f"({', '.join(TARGET_APIS)}) were mentioned.")
        return

    print(f"{len(relevant)} relevant release(s) found:\n")
    for entry in relevant:
        print("\n".join(format_release_block(entry)))


# ============================================================================
# 6. Email (draft file always written when there's something to report;
#    actually sent via SMTP only with --send, same opt-in safety convention
#    phase1_historical_audit.py already uses — send_email.py needs
#    SMTP_HOST/SMTP_USER/SMTP_PASSWORD and fails loudly, not silently, if
#    they're missing)
# ============================================================================

def email_subject(relevant):
    return f"[MDES] {len(relevant)} new release(s) impacting tracked APIs"


def render_email_body(relevant):
    lines = [
        f"{len(relevant)} MDES pre-release note(s) affect the tracked APIs "
        f"({', '.join(TARGET_APIS)}):",
        "",
    ]
    for entry in relevant:
        lines += format_release_block(entry)
    return "\n".join(lines)


def default_report_xlsx_path():
    # Includes time-of-day, not just the date, so re-running the check more
    # than once on the same day writes a NEW file instead of overwriting the
    # previous (possibly still-unsent) report.
    here = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(here, "reports", f"phase2 alert - {datetime.now().strftime('%d-%m-%y %Hh%Mm%S')}.xlsx")


# ============================================================================
# 6bis. Excel report (one sheet per tracked API that was actually mentioned
#       in a relevant release, + a summary sheet) — the only report format,
#       attached to the email. No .md file is written anywhere in this
#       pipeline anymore.
# ============================================================================

HEADER_FILL = PatternFill('solid', fgColor='305496')
HEADER_FONT = Font(color='FFFFFF', bold=True)
TITLE_FONT = Font(bold=True, size=14)
WRAP = Alignment(vertical='top', wrap_text=True)


def _safe_sheet_name(name, used):
    cleaned = re.sub(r'[:\\/?*\[\]]', '', name)[:31].strip() or 'Sheet'
    base, n = cleaned, 2
    while cleaned.lower() in used:
        suffix = f" ({n})"
        cleaned = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(cleaned.lower())
    return cleaned


def _style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_summary_sheet(wb, relevant):
    ws = wb.create_sheet("Résumé", 0)
    ws['A1'] = f"Phase 2 — {len(relevant)} release(s) impactant les APIs suivies"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')

    header_row = 3
    headers = ['Note', 'Version (MDES release)', 'Type', 'Date de mise à jour', 'APIs concernées', 'URL']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(headers))

    r = header_row + 1
    for entry in relevant:
        values = [entry['title'], entry['mdes_release'], entry['note_type'], entry['upgrade_date'],
                   ', '.join(entry['matched_apis']), entry['url']]
        for col, v in enumerate(values, start=1):
            ws.cell(row=r, column=col, value=v).alignment = WRAP
        r += 1

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize(ws, [40, 20, 14, 18, 34, 45])


def _write_api_sheet(wb, sheet_name, api, relevant):
    ws = wb.create_sheet(sheet_name)
    ws['A1'] = api
    ws['A1'].font = TITLE_FONT

    headers = ['Note', 'Version', 'Date de mise à jour', 'Endpoint(s)', 'Changement', 'Description',
               'Champ', 'Type', 'Min', 'Max', 'Required', 'Description du champ']
    header_row = 3
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(headers))

    row = header_row + 1
    for entry in relevant:
        if api not in entry['matched_apis']:
            continue
        changes = entry['parsed']['changes']
        if not changes:
            values = [entry['title'], entry['mdes_release'], entry['upgrade_date'], '', '(aucun changement structuré parsé)',
                       '', '', '', '', '', '', '']
            for col, v in enumerate(values, start=1):
                ws.cell(row=row, column=col, value=v).alignment = WRAP
            row += 1
            continue
        for c in changes:
            endpoints = ', '.join(c['endpoints']) if c['endpoints'] else '(aucun endpoint identifié)'
            if not c['fields']:
                values = [entry['title'], entry['mdes_release'], entry['upgrade_date'], endpoints,
                           c['title'], c.get('description') or '', '', '', '', '', '', '']
                for col, v in enumerate(values, start=1):
                    ws.cell(row=row, column=col, value=v).alignment = WRAP
                row += 1
                continue
            for f in c['fields']:
                values = [
                    entry['title'], entry['mdes_release'], entry['upgrade_date'], endpoints,
                    c['title'], c.get('description') or '', f['name'], f.get('type') or '',
                    f.get('min') or '', f.get('max') or '', f.get('required_raw') or '',
                    f.get('description') or '',
                ]
                for col, v in enumerate(values, start=1):
                    ws.cell(row=row, column=col, value=v).alignment = WRAP
                row += 1

    last_row = row - 1
    if last_row >= header_row + 1:
        table_ref = f"A{header_row}:{get_column_letter(len(headers))}{last_row}"
        table = Table(displayName=f"T{re.sub(r'[^A-Za-z0-9]', '', sheet_name)}", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
        ws.add_table(table)

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize(ws, [34, 16, 16, 26, 34, 40, 30, 12, 8, 8, 12, 40])


def render_report_xlsx(relevant, xlsx_path):
    """Builds the workbook attached to the phase 2 email: a 'Résumé' sheet
    (one row per relevant release) and one sheet per tracked API that was
    actually mentioned in at least one relevant release, each listing every
    change/field concerning that API across all matching releases."""
    wb = Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb, relevant)

    mentioned_apis = [api for api in TARGET_APIS if any(api in e['matched_apis'] for e in relevant)]
    used_names = set()
    for api in mentioned_apis:
        sheet_name = _safe_sheet_name(api, used_names)
        _write_api_sheet(wb, sheet_name, api, relevant)

    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    wb.save(xlsx_path)
    return xlsx_path


# ============================================================================
# 7. CLI
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--cache-dir", default=fetch_mc_source.DEFAULT_CACHE_DIR)
    parser.add_argument("--refresh", action="store_true",
                         help="Re-fetch known_releases.json even if already cached")
    parser.add_argument("--check", choices=["pending", "latest"], default="pending",
                         help="'pending' (default): every release not yet reflected in pre-dig.yaml "
                              "(Production date still in the future) — the authoritative sweep, ADVANCES "
                              "the checkpoint afterward. 'latest': releases newer than whatever 'pending' "
                              "last examined (persisted checkpoint) — does NOT advance the checkpoint, so "
                              "repeated runs keep re-showing the same still-new release(s) until 'pending' "
                              "runs again. Falls back to just the single newest table entry if 'pending' "
                              "has never been run.")
    parser.add_argument("--limit", type=int, default=None,
                         help="Testing only: check the first N releases from the table instead of the "
                              "normal --check selection. Ignored if --all is passed. Checkpoint untouched.")
    parser.add_argument("--all", action="store_true",
                         help="Testing only: check every release currently in the index (e.g. for a "
                              "one-off historical sweep), instead of the normal --check selection. "
                              "Checkpoint untouched.")
    parser.add_argument("--report", default=None,
                         help="Excel report output path (default: reports/phase2 alert - jj-mm-aa HHhMMmSS.xlsx)")
    parser.add_argument("--send", action="store_true",
                         help="Actually email the result via SMTP (needs SMTP_HOST/SMTP_USER/"
                              "SMTP_PASSWORD env vars — see send_email.py, fails loudly if missing). "
                              "Without this flag, only prints to terminal and writes the local .xlsx report.")
    args = parser.parse_args()

    os.makedirs(args.cache_dir, exist_ok=True)

    entries = get_release_notes_table(args.cache_dir, refresh=args.refresh)
    print(f"  [phase2a] {len(entries)} release(s) in the index", file=sys.stderr)

    testing_mode = args.all or args.limit is not None
    if testing_mode:
        limit = None if args.all else args.limit
        print(f"  [phase2a] --all/--limit passed — checking "
              + ("every release" if args.all else f"the latest {limit}")
              + " (checkpoint untouched)", file=sys.stderr)
        relevant = filter_relevant_releases(entries, args.cache_dir, limit=limit)
        print_relevant_releases(relevant)
    elif args.check == "pending":
        result = check_pending_since_predig(entries, args.cache_dir)
        print(f"  [phase2a] check=pending, {len(result['working_set'])} release(s) to check, "
              f"outcome={result['outcome']}", file=sys.stderr)
        print_relevant_releases(result["relevant"])
        advance_checkpoint(args.cache_dir, entries)
        relevant = result["relevant"]
    else:
        result = check_since_checkpoint(entries, args.cache_dir)
        if result["used_fallback"]:
            print("  [phase2a] no checkpoint yet ('pending' has never been run) — "
                  "falling back to just the newest release", file=sys.stderr)
        print(f"  [phase2a] check=latest, {len(result['working_set'])} release(s) to check, "
              f"outcome={result['outcome']}", file=sys.stderr)
        print_relevant_releases(result["relevant"])
        relevant = result["relevant"]

    if not relevant:
        return

    subject = email_subject(relevant)
    body = render_email_body(relevant)

    xlsx_path = args.report or default_report_xlsx_path()
    render_report_xlsx(relevant, xlsx_path)
    print(f"  [phase2a] report (xlsx) -> {xlsx_path}", file=sys.stderr)

    if args.send:
        try:
            sent_to = send_email_module.send_email(subject, body, attachments=[xlsx_path])
            print(f"\n  [phase2a] email SENT to {', '.join(sent_to)} with {os.path.basename(xlsx_path)} attached",
                  file=sys.stderr)
        except send_email_module.SmtpConfigError as e:
            print(f"\n  [phase2a] email NOT sent — {e}", file=sys.stderr)
            sys.exit(1)
    else:
        print(f"\n  [phase2a] xlsx report -> {xlsx_path} (NOT sent — pass --send to actually email it)",
              file=sys.stderr)


if __name__ == "__main__":
    main()
