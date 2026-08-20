#!/usr/bin/env python3
"""
mcs_divergence.py — direct field-by-field diff between Mastercard Checkout
Solutions' official specs (Card API + Cards Batch Enroll API) and data.yaml's
C2P-prefixed implementation.

Same relationship to phase1_historical_audit.py's audit_predig_vs_data_direct()
as that function has to pre-dig.yaml vs data.yaml: same diff engine
(flatten_endpoint_predig / flatten_endpoint / compare_field / find_matches),
reused as-is here, just pointed at a different pair of specs and a different
operation list -- Mastercard Checkout Solutions ships its official spec as two
separate swagger files (Card API, Cards Batch Enroll API) rather than pre-dig's
single pre-dig.yaml, so PRIORITY_OPERATIONS below records which file + path +
HTTP method each tracked data.yaml operation corresponds to.

Only 5 of data.yaml's 8 "*OnC2P"/"CardAddOnService" operations are tracked
here (Enroll/Get/Delete/BulkEnrollment/GetBatchResult) -- those are the ones
matching this specific pair of official specs. UpdateDataOnC2P,
OptInOptOutCardAddOnService and GetCardAddOnServiceStatus don't correspond to
anything on this API's reference page and are left untracked.
"""

import json
import os
import re
import sys
import urllib.request
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

from diff_openapi_all import load_spec
from phase1_historical_audit import (
    flatten_endpoint, flatten_endpoint_predig, compare_field, find_matches,
    MINOR_NOISE_ATTRS, STATUS_LABEL_FR, STATUS_FILL,
    TITLE_FONT, WRAP, _safe_sheet_name, _style_header_row, _autosize,
)

CARD_SPEC_URL = "https://static.developer.mastercard.com/content/mastercard-checkout-solutions/swagger/card_v1.yaml"
BATCH_SPEC_URL = "https://static.developer.mastercard.com/content/mastercard-checkout-solutions/swagger/card_batch_enrollment_scof_v1.yaml"
USER_AGENT = "Mozilla/5.0 (compatible; mc-divergence-bot/1.0)"

DEFAULT_CACHE_DIR = os.path.join(HERE, "cache", "mcs")
CARD_SPEC_PATH = os.path.join(DEFAULT_CACHE_DIR, "card_v1.yaml")
BATCH_SPEC_PATH = os.path.join(DEFAULT_CACHE_DIR, "card_batch_enrollment_scof_v1.yaml")
META_PATH = os.path.join(DEFAULT_CACHE_DIR, "meta.json")
DEFAULT_REPORT_XLSX_PATH = os.path.join(HERE, "reports", "mcs_comparison_report.xlsx")

# (data.yaml path, 'card'|'batch' spec, official path, HTTP method, display name)
PRIORITY_OPERATIONS = [
    ('/EnrollCardOnC2P/V2', 'card', '/cards', 'post', 'Enroll a Card'),
    ('/GetDataFromC2P/V2', 'card', '/cards/{cardId}', 'get', 'Get Card by ID'),
    ('/DeleteDataOnC2P/V2', 'card', '/cards/{cardId}', 'delete', 'Delete Card by ID'),
    ('/BulkEnrollmentOnC2P/V2', 'batch', '/cards/batch', 'post', 'Batch Cards Enrollment'),
    ('/GetC2PBatchResult/V2', 'batch', '/cards/batch/{batchId}', 'get', 'Get Batch Status'),
]


def _fetch(url):
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read().decode("utf-8")


def fetch_specs(cache_dir=DEFAULT_CACHE_DIR):
    os.makedirs(cache_dir, exist_ok=True)
    for url, path in ((CARD_SPEC_URL, CARD_SPEC_PATH), (BATCH_SPEC_URL, BATCH_SPEC_PATH)):
        text = _fetch(url)
        with open(path, "w", encoding="utf-8") as f:
            f.write(text)
    with open(META_PATH, "w", encoding="utf-8") as f:
        json.dump({
            "card_spec_url": CARD_SPEC_URL, "batch_spec_url": BATCH_SPEC_URL,
            "fetched_at": datetime.now(timezone.utc).isoformat(),
        }, f, indent=2)


def audit_mcs_vs_data_direct(card_spec, batch_spec, data_spec):
    """Returns {data_yaml_path: {'display', 'endpoint_exists_in_data',
    'endpoint_exists_in_mcs', 'fields': [...]}} -- same per-field shape as
    audit_predig_vs_data_direct() (type/minLength/maxLength/required/
    description kept on every result regardless of status)."""
    specs = {'card': card_spec, 'batch': batch_spec}
    results = {}
    for data_path, spec_key, mc_path, method, display in PRIORITY_OPERATIONS:
        mc_fields = flatten_endpoint_predig(specs[spec_key], mc_path, method=method)
        data_fields = flatten_endpoint(data_spec, data_path, method='post')

        if mc_fields is None:
            results[data_path] = {
                'display': display, 'endpoint_exists_in_data': data_fields is not None,
                'endpoint_exists_in_mcs': False, 'fields': [],
            }
            continue

        field_results = []
        for key, info in mc_fields.items():
            leaf_name = re.split(r'[.\[\]]+', key)[-1]
            expected = {
                'type': info.get('type'), 'minLength': info.get('minLength'),
                'maxLength': info.get('maxLength'), 'required': info.get('required'),
                'description': info.get('description'),
            }
            if data_fields is None:
                field_results.append({'name': key, 'status': 'non_implemente', 'reliable': True,
                                       'reasons': ["endpoint missing from data.yaml"], **expected})
                continue
            matches = find_matches(leaf_name, data_fields)
            if not matches:
                field_results.append({'name': key, 'status': 'non_implemente', 'reliable': True,
                                       'reasons': ["no matching field found"], **expected})
                continue
            best = next((m for m in matches if m[1]), matches[0])
            pseudo = {'name': leaf_name, 'type': info.get('type'), 'min': info.get('minLength'),
                      'max': info.get('maxLength'), 'required': info.get('required')}
            status, reasons, mismatched = compare_field(pseudo, data_fields[best[0]])
            if not best[1]:
                reasons = reasons + ["name found but nesting differs — needs review"]
            reliable = not mismatched or not mismatched.issubset(MINOR_NOISE_ATTRS)
            field_results.append({'name': key, 'status': status, 'reasons': reasons,
                                   'reliable': reliable, **expected})

        results[data_path] = {
            'display': display, 'endpoint_exists_in_data': data_fields is not None,
            'endpoint_exists_in_mcs': True, 'fields': field_results,
        }
    return results


def render_report_xlsx(mcs_direct, xlsx_path):
    wb = Workbook()
    wb.remove(wb.active)

    used_names = set()
    op_sheet_names = {}
    for data_path, _, _, _, display in PRIORITY_OPERATIONS:
        op_sheet_names[data_path] = _safe_sheet_name(display, used_names)

    _write_summary_sheet(wb, mcs_direct, op_sheet_names)
    for data_path, _, _, _, display in PRIORITY_OPERATIONS:
        entry = mcs_direct.get(data_path, {'display': display})
        _write_operation_sheet(wb, op_sheet_names[data_path], data_path, entry)

    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    wb.save(xlsx_path)
    return xlsx_path


def _summarize(entry):
    if not entry.get('endpoint_exists_in_mcs'):
        return {'status': 'unknown', 'reliable_issues': 0, 'total_fields': 0}
    if not entry['endpoint_exists_in_data']:
        return {'status': 'non_implemente', 'reliable_issues': len(entry['fields']), 'total_fields': len(entry['fields'])}
    reliable_problems = [f for f in entry['fields'] if f['status'] != 'implemente' and f['reliable']]
    if any(f['status'] == 'non_implemente' for f in reliable_problems):
        status = 'non_implemente'
    elif reliable_problems:
        status = 'partiel'
    else:
        status = 'implemente'
    return {'status': status, 'reliable_issues': len(reliable_problems), 'total_fields': len(entry['fields'])}


def _write_summary_sheet(wb, mcs_direct, op_sheet_names):
    ws = wb.create_sheet("Summary", 0)
    ws['A1'] = "Mastercard Checkout Solutions — data.yaml vs official spec"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')

    header_row = 3
    headers = ['Operation', 'data.yaml path', 'Status', 'Reliable gaps', 'Total fields', 'Detail']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(headers))

    r = header_row + 1
    for data_path, _, _, _, display in PRIORITY_OPERATIONS:
        entry = mcs_direct.get(data_path, {})
        summary = _summarize(entry)
        ws.cell(row=r, column=1, value=display)
        ws.cell(row=r, column=2, value=data_path)
        status_cell = ws.cell(row=r, column=3, value=STATUS_LABEL_FR.get(summary['status'], summary['status']))
        status_cell.fill = STATUS_FILL.get(summary['status'])
        ws.cell(row=r, column=4, value=summary['reliable_issues'])
        ws.cell(row=r, column=5, value=summary['total_fields'])
        link_cell = ws.cell(row=r, column=6, value=f"See '{op_sheet_names[data_path]}'")
        link_cell.hyperlink = f"#'{op_sheet_names[data_path]}'!A1"
        link_cell.font = Font(color='0563C1', underline='single')
        r += 1

    _autosize(ws, [30, 26, 16, 14, 13, 24])


def _write_operation_sheet(wb, sheet_name, data_path, entry):
    ws = wb.create_sheet(sheet_name)
    ws['A1'] = entry['display']
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"data.yaml path: {data_path}"
    ws['A2'].font = Font(italic=True)

    if not entry.get('endpoint_exists_in_mcs'):
        ws['A4'] = "Endpoint not found in the official Mastercard Checkout Solutions spec."
        _autosize(ws, [60])
        return
    if not entry['endpoint_exists_in_data']:
        ws['A4'] = f"Endpoint missing from data.yaml — {len(entry['fields'])} field(s) not verifiable."
        ws['A4'].font = Font(bold=True, color='9C0006')

    headers = ['Field', 'Status', 'Reliable', 'Reasons', 'Expected type', 'MinLength', 'MaxLength', 'Required', 'Description']
    header_row = 6
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(headers))

    fields_sorted = sorted(entry['fields'], key=lambda f: (f['status'] == 'implemente', not f['reliable'], f['name']))
    row = header_row + 1
    for f in fields_sorted:
        values = [
            f['name'], STATUS_LABEL_FR[f['status']], 'Yes' if f['reliable'] else 'No',
            '; '.join(f['reasons']), f.get('type') or '',
            f.get('minLength') if f.get('minLength') is not None else '',
            f.get('maxLength') if f.get('maxLength') is not None else '',
            '' if f.get('required') is None else ('Yes' if f['required'] else 'No'),
            f.get('description') or '',
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.alignment = WRAP
            cell.fill = STATUS_FILL[f['status']]
        row += 1

    last_row = row - 1
    if last_row >= header_row + 1:
        table_ref = f"A{header_row}:{get_column_letter(len(headers))}{last_row}"
        table = Table(displayName=f"T{re.sub(r'[^A-Za-z0-9]', '', sheet_name)}", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
        ws.add_table(table)

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize(ws, [38, 16, 9, 45, 14, 11, 11, 10, 40])


def render_email_body(mcs_direct):
    total_missing = sum(len([f for f in e.get('fields', []) if f['status'] != 'implemente' and f['reliable']])
                         for e in mcs_direct.values())
    lines = [
        "Hello,", "",
        "The Mastercard Checkout Solutions audit (official spec vs data.yaml) has been run. "
        "Summary below, field-by-field detail in the attachment.",
        "",
    ]
    for data_path, _, _, _, display in PRIORITY_OPERATIONS:
        entry = mcs_direct.get(data_path, {})
        summary = _summarize(entry)
        lines.append(f"- **{display}** ({data_path}): {summary['reliable_issues']} gap(s) "
                     f"on {summary['total_fields']} field(s)")
    lines.append('')
    return total_missing, '\n'.join(lines)
