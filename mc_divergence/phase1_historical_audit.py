#!/usr/bin/env python3
"""
phase1_historical_audit.py — Phase 1 (batch, one-shot) of the Mastercard MDES
divergence pipeline described in mastercard_divergence_prompt.md.

Pipeline:
  1. Load the release-notes table (cache/known_releases.json, refreshed via
     fetch_mc_source.py) and keep entries from CUTOFF (default: Jan 2025)
     to today.
  2. For each kept release, fetch (or reuse the cached copy of) its note and
     parse it with parse_release_note.py -> list of {title, endpoints, fields}.
  3. For every field of every change, check it against the CURRENT PowerCARD
     data.yaml using diff_openapi_all.py's $ref-resolving engine (load_spec /
     get_operation / extract_request_schema / extract_response_schemas /
     flatten_schema — already built and reused, not reimplemented):
       - leaf not found anywhere in that endpoint's schema  -> non_implemente
       - found, but type/min/max/required disagree          -> partiel
       - found and every comparable attribute agrees         -> implemente
     (matching is leaf-name + parent-chain based, same heuristic validated
     earlier in check_field_changes.py against pre-dig.yaml vs data.yaml —
     folds case and singular/plural, and 'cypher'/'cipher' spelling)
  4. Write a Markdown report (global adoption rate + per-release breakdown,
     "partiel"/"non_implemente" surfaced first, "implemente" not left out
     but pushed to the bottom so it doesn't drown the real gaps) and a
     separate email-draft Markdown file (no SMTP send — per earlier
     decision, this stays a local file until the mail transport is chosen).

Pre-dig.yaml (Mastercard's own current spec, used as the authoritative side
of the direct structural diff — see audit_predig_vs_data_direct()) is always
re-fetched fresh from Mastercard at the start of every run, not read from
whatever happened to be cached last time — so a new Mastercard release is
picked up automatically without ever needing --refresh or a manual re-download.

Usage:
    python3 phase1_historical_audit.py
        [--data-yaml ../data.yaml] [--cutoff 2025-01] [--refresh]
        [--report reports/phase1_divergence_report.md]
        [--email reports/phase1_email_draft.md]
"""

import argparse
import json
import os
import re
import sys
import urllib.request

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")
from datetime import date, datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from diff_openapi_all import (
    load_spec, get_operation, extract_request_schema,
    extract_response_schemas, flatten_schema, resolve_refs,
)
from parse_release_note import parse_note, DEFAULT_CACHE_DIR as NOTES_CACHE_DIR
import fetch_mc_source
import send_email as send_email_module

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_DATA_YAML = os.environ.get('DATA_YAML_PATH', r'C:\Users\moham\Desktop\input\data.yaml')
DEFAULT_REPORT_XLSX_PATH = os.path.join(HERE, "reports", "phase1_divergence_report.xlsx")

# Recipients not decided yet (HPS infra / SMTP vs Graph still to confirm) —
# no real send wired up by default; --send opts in explicitly.
EMAIL_RECIPIENTS = os.environ.get('MC_DIVERGENCE_RECIPIENTS')  # comma-separated, unset = TBD


# The 7 APIs the pipeline prioritizes end-to-end (prompt: "APIs prioritaires").
PRIORITY_PATHS = {
    '/requestactivationmethods', '/deliveractivationcode',
    '/authorizeservice', '/notifyserviceactivated', '/notifytokenupdated',
    '/validateactivationcode', '/getaccountinformation',
}


# ============================================================================
# 1. Release-notes selection (cutoff filter on the "YYYY Month" mdes_release
#    field, which is present and consistently formatted even on the oldest
#    2020-era entries)
# ============================================================================

MONTH_NUM = {
    'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
    'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7,
    'aug': 8, 'august': 8, 'sep': 9, 'sept': 9, 'september': 9,
    'oct': 10, 'october': 10, 'nov': 11, 'november': 11, 'dec': 12, 'december': 12,
}
MDES_RELEASE_RE = re.compile(r'^(\d{4})\s+([A-Za-z]+)')


def parse_mdes_release(mdes_release):
    m = MDES_RELEASE_RE.match(mdes_release or '')
    if not m:
        return None
    year = int(m.group(1))
    month = MONTH_NUM.get(m.group(2).strip().lower())
    if month is None:
        return None
    return (year, month)


def parse_cutoff(cutoff_str):
    """'2025-01' -> (2025, 1)"""
    y, m = cutoff_str.split('-')
    return (int(y), int(m))


def load_releases_since(cache_dir, cutoff):
    path = os.path.join(cache_dir, 'known_releases.json')
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    kept = []
    for entry in data['releases']:
        ym = parse_mdes_release(entry['mdes_release'])
        if ym is not None and ym >= cutoff:
            kept.append(entry)
    kept.sort(key=lambda e: parse_mdes_release(e['mdes_release']))
    return kept


# ============================================================================
# 2. Note fetching (reuse cache/notes_raw/<slug>.md if present, else pull)
# ============================================================================

def slug_from_url(url):
    return url.rstrip('/').split('/')[-2] if url.rstrip('/').endswith('index.md') else \
        url.rstrip('/').split('/')[-1]


def get_note_markdown(url, cache_dir):
    slug = slug_from_url(url)
    raw_dir = os.path.join(cache_dir, 'notes_raw')
    os.makedirs(raw_dir, exist_ok=True)
    path = os.path.join(raw_dir, f"{slug}.md")
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return f.read(), path
    req = urllib.request.Request(url, headers={"User-Agent": fetch_mc_source.USER_AGENT})
    with urllib.request.urlopen(req, timeout=30) as resp:
        text = resp.read().decode('utf-8')
    with open(path, 'w', encoding='utf-8') as f:
        f.write(text)
    return text, path


# ============================================================================
# 3. Field matching against data.yaml (same heuristic as check_field_changes.py)
# ============================================================================

def normalize_segment(seg):
    s = re.sub(r'[^a-z0-9]', '', seg.lower())
    return s.replace('cypher', 'cipher')


def leaf_equal(a, b):
    return a == b or a.rstrip('s') == b.rstrip('s')


def find_matches(field_name, flat_fields):
    """field_name may use '.' (e.g. 'activationMethod.type') or be a bare
    leaf. Returns [(full_key, chain_ok), ...]."""
    segments = [s.strip() for s in re.split(r'->|\.', field_name) if s.strip()]
    norm_segments = [normalize_segment(s) for s in segments]
    leaf = norm_segments[-1]
    parents = norm_segments[:-1]

    matches = []
    for key in flat_fields:
        key_segments = [s for s in re.split(r'[.\[\]]+', key) if s]
        norm_key_segments = [normalize_segment(s) for s in key_segments]
        if not norm_key_segments or not leaf_equal(norm_key_segments[-1], leaf):
            continue
        chain_ok = True
        ptr = 0
        for p in parents:
            found_at = None
            for i in range(ptr, len(norm_key_segments) - 1):
                if p in norm_key_segments[i] or norm_key_segments[i] in p:
                    found_at = i
                    break
            if found_at is None:
                chain_ok = False
                break
            ptr = found_at + 1
        matches.append((key, chain_ok))
    return matches


# note "Data Type" strings -> the handful of JSON-Schema 'type' values they
# correspond to, for the (best-effort, non-blocking) type comparison
NOTE_TYPE_TO_SCHEMA_TYPE = {
    'string': 'string', 'number': 'integer', 'object': 'object',
    'complexobject': 'object', 'array': 'array', 'arrayofstring': 'array',
    'arraystring': 'array',
}


#  Attributes whose mismatch is likely noise rather than a real gap: for the
#  5 priority endpoints, data.yaml systematically has minLength=0 and no
#  `required` list on most legacy-style schemas (a generation-tool
#  convention, confirmed by sampling — 'data.yaml a 0'/'a False' on ~100% of
#  cases), not a genuine documentation omission per field. 'type'/'maxLength'
#  mismatches don't show that pattern and stay reliable.
MINOR_NOISE_ATTRS = {'minLength', 'required'}


def compare_field(note_field, matched_info):
    """Returns (status, reasons[], mismatched_attrs:set) for ONE matched
    field, status in {'implemente', 'partiel'}. Only flags a mismatch on
    attributes the note actually specifies — a note that omits
    type/min/max/required gives no grounds to downgrade, so an unremarked
    field with a name match stays 'implemente'."""
    reasons = []
    mismatched_attrs = set()

    note_type = note_field.get('type')
    if note_type:
        norm_note_type = normalize_segment(note_type)
        expected = NOTE_TYPE_TO_SCHEMA_TYPE.get(norm_note_type)
        actual = matched_info.get('type')
        if expected and actual and expected != actual:
            reasons.append(f"type: note dit '{note_type}' ({expected}), data.yaml a '{actual}'")
            mismatched_attrs.add('type')

    for note_key, schema_key, label in (('min', 'minLength', 'minLength'), ('max', 'maxLength', 'maxLength')):
        note_val = note_field.get(note_key)
        if note_val and re.match(r'^\d+$', str(note_val).strip()):
            actual = matched_info.get(schema_key)
            if actual is not None and int(note_val) != int(actual):
                reasons.append(f"{label}: note dit {note_val}, data.yaml a {actual}")
                mismatched_attrs.add(schema_key)

    if note_field.get('required') is not None:
        actual_required = bool(matched_info.get('required'))
        if note_field['required'] != actual_required:
            reasons.append(
                f"required: note dit {note_field['required']}, data.yaml a {actual_required}"
            )
            mismatched_attrs.add('required')

    return ('partiel' if reasons else 'implemente'), reasons, mismatched_attrs


def flatten_endpoint(spec, path, method='post'):
    op = get_operation(spec, path, method)
    if op is None:
        return None  # endpoint doesn't exist at all in data.yaml
    fields = {}
    req_schema, _ = extract_request_schema(op, spec)
    if req_schema:
        for k, v in flatten_schema(req_schema).items():
            fields[f'request.{k}'] = v
    for code, schema in extract_response_schemas(op, spec).items():
        if schema:
            for k, v in flatten_schema(schema).items():
                fields[f'response[{code}].{k}'] = v
    return fields


def endpoint_ref_to_path(raw_endpoint):
    """'GET /authorizeService' or '(inferred) /authorizeService' -> '/authorizeService'"""
    m = re.search(r'(/\S+)$', raw_endpoint)
    return m.group(1) if m else raw_endpoint


def flatten_endpoint_predig(spec, path, method='post'):
    """Same as flatten_endpoint(), but resolves requestBody/response as a
    whole BEFORE reading '.content' out of them. Needed only for
    pre-dig.yaml: unlike data.yaml (which always inlines requestBody/
    response objects and only $refs at the schema level), Mastercard's own
    spec points requestBody/each response at a whole-object $ref into
    components/requestBodies / components/responses (confirmed earlier by
    diff_predig_vs_data.py). Skipping this step silently produces "endpoint
    has 0 fields" instead of a real diff."""
    op = get_operation(spec, path, method)
    if op is None:
        return None
    resolved_op = dict(op)
    if op.get('requestBody'):
        resolved_op['requestBody'] = resolve_refs(op['requestBody'], spec)
    if op.get('responses'):
        resolved_op['responses'] = {
            code: resolve_refs(resp, spec) for code, resp in op['responses'].items()
        }
    fields = {}
    req_schema, _ = extract_request_schema(resolved_op, spec)
    if req_schema:
        for k, v in flatten_schema(req_schema).items():
            fields[f'request.{k}'] = v
    for code, schema in extract_response_schemas(resolved_op, spec).items():
        if schema:
            for k, v in flatten_schema(schema).items():
                fields[f'response[{code}].{k}'] = v
    return fields


# ============================================================================
# 3bis0. Schema-provenance tagging — which NAMED schema component (the
#        $ref target, e.g. 'chipDataSchema') a field was actually defined
#        in, on both sides. Two endpoints whose missing/divergent field
#        traces back to the identical component name on BOTH pre-dig.yaml
#        AND data.yaml share one root cause: fixing that single shared
#        schema resolves the gap for every endpoint that $refs it, instead
#        of needing a per-endpoint fix (see group_shared_schema_fixes()
#        below). Purely mechanical — string comparison of $ref targets,
#        no LLM involved.
# ============================================================================

def _resolve_ref_hop(node, root):
    """Resolve a bare {'$ref': ...} pointer ONE hop, without recursing into
    the target's own nested $refs — used to walk wrapper structures
    (requestBody/response -> content -> schema) without prematurely
    resolving the schema subtree itself, so origin tagging further down
    can still see the $ref names inside it. Returns None on a dangling ref."""
    if not (isinstance(node, dict) and isinstance(node.get('$ref'), str) and node['$ref'].startswith('#/')):
        return node
    target = root
    try:
        for part in node['$ref'][2:].split('/'):
            target = target[part]
    except (KeyError, TypeError):
        return None
    return target


def _raw_operation_schemas(op, root):
    """Like extract_request_schema()/extract_response_schemas(), but only
    hops through wrapper-level $refs (requestBody/response -> content ->
    schema) instead of fully resolving the schema subtree — so the schema
    node handed back may itself still be (or contain) raw $refs, which
    flatten_schema_with_origin() then walks lazily while tagging provenance.
    Works for both pre-dig.yaml (requestBody/response are whole-object
    $refs into components/requestBodies|responses) and data.yaml (always
    inlined, $ref only at the schema level) — the same hop-if-ref logic
    handles both shapes."""
    req_schema = None
    rb = _resolve_ref_hop(op.get('requestBody'), root)
    if rb:
        content = (rb or {}).get('content', {}) or {}
        entry = content.get('application/json') or next(iter(content.values()), None)
        entry = _resolve_ref_hop(entry, root)
        req_schema = (entry or {}).get('schema')

    resp_schemas = {}
    for code, resp in (op.get('responses') or {}).items():
        resp = _resolve_ref_hop(resp, root)
        content = (resp or {}).get('content', {}) or {}
        entry = content.get('application/json') or next(iter(content.values()), None)
        entry = _resolve_ref_hop(entry, root)
        resp_schemas[code] = (entry or {}).get('schema') if entry else None
    return req_schema, resp_schemas


def flatten_schema_with_origin(schema, root, prefix="", origin=None, seen=None, depth=0):
    """Mirrors flatten_schema()'s path construction, but resolves $refs
    itself (rather than requiring an already-fully-resolved schema) so it
    can tag each field with the name of the nearest named schema component
    it was reached through. Returns {field_path: origin_schema_name}."""
    if seen is None:
        seen = set()
    origins = {}
    if not isinstance(schema, dict) or depth > 25:
        return origins

    ref = schema.get('$ref')
    if isinstance(ref, str) and ref.startswith('#/'):
        if ref in seen:
            return origins
        target = _resolve_ref_hop(schema, root)
        if target is None:
            return origins
        ref_name = ref.rsplit('/', 1)[-1]
        return flatten_schema_with_origin(target, root, prefix, origin=ref_name,
                                           seen=seen | {ref}, depth=depth + 1)

    props = schema.get('properties', {}) or {}
    for name, sub in props.items():
        full = f"{prefix}{name}"
        if not isinstance(sub, dict):
            continue
        origins[full] = origin
        if '$ref' in sub or sub.get('type') == 'object' or 'properties' in sub:
            origins.update(flatten_schema_with_origin(sub, root, full + ".", origin=origin,
                                                        seen=seen, depth=depth + 1))
        elif sub.get('type') == 'array' and isinstance(sub.get('items'), dict) and \
                ('$ref' in sub['items'] or 'properties' in sub['items']):
            origins.update(flatten_schema_with_origin(sub['items'], root, full + "[].", origin=origin,
                                                        seen=seen, depth=depth + 1))
    return origins


def endpoint_field_origins(spec, path, method='post'):
    """{field_path: origin_schema_name} for one endpoint — same key format
    ('request.x.y', 'response[200].x.y') as flatten_endpoint()/
    flatten_endpoint_predig(), so it can be looked up with the exact same
    keys already used in audit_predig_vs_data_direct()."""
    op = get_operation(spec, path, method)
    if op is None:
        return {}
    req_schema, resp_schemas = _raw_operation_schemas(op, spec)
    origins = {}
    if req_schema:
        for k, v in flatten_schema_with_origin(req_schema, spec).items():
            origins[f'request.{k}'] = v
    for code, schema in resp_schemas.items():
        if schema:
            for k, v in flatten_schema_with_origin(schema, spec).items():
                origins[f'response[{code}].{k}'] = v
    return origins


def closest_origin(path, origins):
    """Walk a dotted field path up toward its root until a known origin is
    found — used when a field has no match at all on the other side, so the
    nearest ancestor's owning schema stands in as 'where the fix would have
    to go' (e.g. a missing 'chipData.iccSystemRelatedData' leaf falls back
    to whichever schema owns 'encryptedData', its closest existing parent)."""
    cur = path
    while cur:
        if cur in origins:
            return origins[cur]
        if '.' not in cur:
            return None
        cur = cur.rsplit('.', 1)[0]
    return None


def group_shared_schema_fixes(predig_direct):
    """Clusters reliable non-implemented/partial fields that share the
    EXACT same (field path, pre-dig.yaml origin schema, data.yaml origin
    schema, status) across 2+ endpoints — i.e. the field is defined once in
    a shared pre-dig.yaml component and blocked by the same shared
    data.yaml component everywhere it's used. Fixing that one data.yaml
    schema resolves every endpoint in the cluster at once. Endpoints whose
    matching schema differs (even if the leaf name coincidentally matches)
    are NOT clustered, since predig_origin/data_origin wouldn't agree.
    Returns clusters sorted by endpoint count, descending."""
    buckets = {}
    for path, entry in predig_direct.items():
        if not entry.get('endpoint_exists_in_predig') or not entry.get('endpoint_exists_in_data'):
            continue
        for f in entry['fields']:
            if f['status'] == 'implemente' or not f['reliable']:
                continue
            predig_origin = f.get('predig_origin')
            data_origin = f.get('data_origin')
            if not predig_origin:
                continue
            key = (f['name'], predig_origin, data_origin, f['status'])
            bucket = buckets.setdefault(key, {'endpoints': [], 'predig_line': f.get('predig_line'),
                                               'data_line': f.get('data_line')})
            bucket['endpoints'].append((path, entry['display']))

    clusters = []
    for (name, predig_origin, data_origin, status), bucket in buckets.items():
        if len(bucket['endpoints']) < 2:
            continue
        clusters.append({
            'field': name, 'predig_origin': predig_origin, 'data_origin': data_origin,
            'status': status, 'endpoints': bucket['endpoints'],
            'predig_line': bucket['predig_line'], 'data_line': bucket['data_line'],
        })
    clusters.sort(key=lambda c: len(c['endpoints']), reverse=True)
    return clusters


def summarize_schema_fix_clusters(field_clusters):
    """Rolls the per-field clusters from group_shared_schema_fixes() up to
    the (predig_origin, data_origin) schema-pair level — the actually
    actionable unit ('fix data.yaml's EncryptedPayload schema once'),
    instead of one line per field. Merges every field/endpoint that shares
    the same schema pair, regardless of status, so 'N champs sur M
    endpoints' reflects the true blast radius of one schema edit. Each
    field keeps its own source line in both files (a leaf's own key line
    within the shared schema, not the schema's opening line), so the team
    can jump straight to the exact spot to fix — same schema pair, but
    'source' and 'chipData' don't sit on the same line."""
    rollup = {}
    for c in field_clusters:
        key = (c['predig_origin'], c['data_origin'])
        r = rollup.setdefault(key, {'predig_origin': c['predig_origin'], 'data_origin': c['data_origin'],
                                     'fields': {}, 'endpoints': set()})
        r['fields'][c['field']] = {'predig_line': c['predig_line'], 'data_line': c['data_line']}
        r['endpoints'].update(disp for _, disp in c['endpoints'])

    out = []
    for r in rollup.values():
        fields = [{'name': name, **info} for name, info in sorted(r['fields'].items())]
        out.append({
            'predig_origin': r['predig_origin'], 'data_origin': r['data_origin'],
            'field_count': len(fields), 'fields': fields,
            'endpoint_count': len(r['endpoints']), 'endpoints': sorted(r['endpoints']),
        })
    out.sort(key=lambda r: (r['endpoint_count'], r['field_count']), reverse=True)
    return out


def summarize_by_data_target(field_clusters):
    """Rolls schema-pair clusters up one level further, by data.yaml target
    alone — merging every pre-dig.yaml sub-schema that funnels into the same
    data.yaml component. This is the single most actionable headline: 'fix
    data.yaml's EncryptedPayload schema once' covers everything reported
    under it, regardless of how many different pre-dig.yaml components
    (fundingAccountDataSchema, chipDataSchema, accountHolderDataSchema, ...)
    happen to define the individual fields."""
    rollup = {}
    for c in field_clusters:
        if not c['data_origin']:
            continue
        r = rollup.setdefault(c['data_origin'], {'data_origin': c['data_origin'], 'fields': set(),
                                                   'endpoints': set(), 'predig_origins': set()})
        r['fields'].add(c['field'])
        r['endpoints'].update(disp for _, disp in c['endpoints'])
        r['predig_origins'].add(c['predig_origin'])

    out = []
    for r in rollup.values():
        out.append({
            'data_origin': r['data_origin'], 'field_count': len(r['fields']),
            'endpoint_count': len(r['endpoints']), 'endpoints': sorted(r['endpoints']),
            'predig_origin_count': len(r['predig_origins']), 'predig_origins': sorted(r['predig_origins']),
        })
    out.sort(key=lambda r: (r['endpoint_count'], r['field_count']), reverse=True)
    return out


# ============================================================================
# 3bis. Direct structural diff: pre-dig.yaml (live Mastercard spec, already
#       reflects every PAST release note) vs data.yaml, on the 5 priority
#       endpoints. This is the authoritative check — no note-text-parsing
#       fragility involved, every field pre-dig.yaml documents gets a real
#       verdict. The release-notes-based audit below stays useful for WHEN/
#       WHY context (which note introduced a field) and for changes whose
#       Production date hasn't happened yet (so naturally not in pre-dig.yaml
#       yet — not a real gap, just "upcoming").
# ============================================================================

PRIORITY_PATH_DISPLAY = [
    ('/requestActivationMethods', 'Request Activation Methods (RAM)'),
    ('/deliverActivationCode', 'Deliver Activation Code (DAC)'),
    ('/authorizeService', 'Authorize Service (AS)'),
    ('/notifyServiceActivated', 'Notify Service Activated (NSA)'),
    ('/notifyTokenUpdated', 'Notify Token Updated (NTU)'),
    ('/validateActivationCode', 'Validate Activation Code (VAC)'),
    ('/getAccountInformation', 'Get Account Information (GAI)'),
]


def audit_predig_vs_data_direct(predig_spec, data_spec):
    """Returns {path: {'display': str, 'endpoint_exists_in_data': bool,
    'fields': [{'name','status','reasons','reliable','type','minLength',
    'maxLength','required','description'}]}} for the 5 priority endpoints,
    comparing EVERY field pre-dig.yaml documents against data.yaml — not
    just the ones a release note happened to mention. type/minLength/
    maxLength/required/description are pre-dig.yaml's OWN values for that
    field (the expected/Mastercard-documented shape), kept on every result
    regardless of status so a detail view can show the full field table,
    not just the mismatches."""
    results = {}
    for path, display in PRIORITY_PATH_DISPLAY:
        predig_fields = flatten_endpoint_predig(predig_spec, path)
        data_fields = flatten_endpoint(data_spec, path)

        if predig_fields is None:
            results[path] = {'display': display, 'endpoint_exists_in_data': data_fields is not None,
                              'endpoint_exists_in_predig': False, 'fields': []}
            continue

        predig_origins = endpoint_field_origins(predig_spec, path)
        data_origins = endpoint_field_origins(data_spec, path) if data_fields is not None else {}

        field_results = []
        for key, info in predig_fields.items():
            leaf_name = re.split(r'[.\[\]]+', key)[-1]
            predig_origin = predig_origins.get(key)
            expected = {
                'type': info.get('type'), 'minLength': info.get('minLength'),
                'maxLength': info.get('maxLength'), 'required': info.get('required'),
                'description': info.get('description'),
            }
            predig_line = info.get('line')
            if data_fields is None:
                field_results.append({'name': key, 'status': 'non_implemente',
                                       'reasons': ["endpoint absent de data.yaml"], 'reliable': True,
                                       'predig_origin': predig_origin, 'data_origin': None,
                                       'predig_line': predig_line, 'data_line': None, **expected})
                continue
            matches = find_matches(leaf_name, data_fields)
            if not matches:
                field_results.append({'name': key, 'status': 'non_implemente',
                                       'reasons': ["aucun champ correspondant trouvé"], 'reliable': True,
                                       'predig_origin': predig_origin,
                                       'data_origin': closest_origin(key, data_origins),
                                       'predig_line': predig_line, 'data_line': None, **expected})
                continue
            best = next((m for m in matches if m[1]), matches[0])
            pseudo_note_field = {
                'name': leaf_name, 'type': info.get('type'),
                'min': info.get('minLength'), 'max': info.get('maxLength'),
                'required': info.get('required'),
            }
            status, reasons, mismatched_attrs = compare_field(pseudo_note_field, data_fields[best[0]])
            if not best[1]:
                reasons = reasons + ["nom trouvé mais imbrication différente — à vérifier"]
            # 'reliable' = False only when EVERY mismatched attribute is one
            # of the known-noisy ones (minLength/required) — a field that
            # also mismatches on type/maxLength stays reliable even if it
            # additionally shows a noisy attribute.
            reliable = not mismatched_attrs or not mismatched_attrs.issubset(MINOR_NOISE_ATTRS)
            field_results.append({'name': key, 'status': status, 'reasons': reasons, 'reliable': reliable,
                                   'predig_origin': predig_origin, 'data_origin': data_origins.get(best[0]),
                                   'predig_line': predig_line, 'data_line': data_fields[best[0]].get('line'),
                                   **expected})

        results[path] = {'display': display, 'endpoint_exists_in_data': data_fields is not None,
                          'endpoint_exists_in_predig': True, 'fields': field_results}
    return results


# ============================================================================
# 4. Auditing one change against data.yaml
# ============================================================================

NO_SPEC_UPDATE_RE = re.compile(r'(?i)no update to the api spec')


def audit_change(change, data_spec, endpoint_cache):
    """Returns a list of per-endpoint results:
    [{'endpoint': path, 'endpoint_exists': bool, 'fields': [
        {'name':, 'status':, 'matched_path':, 'reasons': []}
    ], 'status': worst-of-fields}]"""
    results = []
    endpoints = change['endpoints'] or ['(unknown) /??']
    for raw_ep in endpoints:
        path = endpoint_ref_to_path(raw_ep)
        if path not in endpoint_cache:
            endpoint_cache[path] = flatten_endpoint(data_spec, path)
        flat_fields = endpoint_cache[path]

        if flat_fields is None:
            results.append({
                'endpoint': path, 'endpoint_exists': False,
                'fields': [], 'status': 'non_implemente',
            })
            continue

        field_results = []
        for nf in change['fields']:
            matches = find_matches(nf['name'], flat_fields)
            if not matches:
                field_results.append({
                    'name': nf['name'], 'status': 'non_implemente',
                    'matched_path': None, 'reasons': ["aucun champ correspondant trouvé"],
                })
                continue
            best = next((m for m in matches if m[1]), matches[0])
            status, reasons, _mismatched = compare_field(nf, flat_fields[best[0]])
            if not best[1]:
                reasons = reasons + ["nom trouvé mais imbrication différente — à vérifier"]
            field_results.append({
                'name': nf['name'], 'status': status,
                'matched_path': best[0], 'reasons': reasons,
            })

        if not field_results:
            # The note itself gave no structured fields to check (prose /
            # bare bullet list, no table — see parse_release_note.py's
            # known limitation) — nothing was actually verified, so don't
            # claim 'implemente'. Exception: the note explicitly says there
            # is no spec change (e.g. "no update to the API specification"),
            # in which case there is genuinely nothing to check.
            explicitly_no_change = bool(NO_SPEC_UPDATE_RE.search(change.get('description') or ''))
            if not change['fields'] and not explicitly_no_change:
                ep_status = 'a_verifier_manuellement'
            else:
                ep_status = 'implemente' if flat_fields else 'non_implemente'
        elif any(f['status'] == 'non_implemente' for f in field_results):
            ep_status = 'non_implemente'
        elif any(f['status'] == 'partiel' for f in field_results):
            ep_status = 'partiel'
        else:
            ep_status = 'implemente'

        results.append({
            'endpoint': path, 'endpoint_exists': True,
            'fields': field_results, 'status': ep_status,
        })
    return results


STATUS_RANK = {'non_implemente': 3, 'a_verifier_manuellement': 2, 'partiel': 1, 'implemente': 0}


def worst_status(statuses):
    return max(statuses, key=lambda s: STATUS_RANK[s]) if statuses else 'implemente'


# ============================================================================
# 5. (former Markdown report renderers removed — the Excel workbook built by
#    render_report_xlsx() below, one sheet per priority API, is now the only
#    report format: easier for the dev team to read than the old .md.)
# ============================================================================



# ============================================================================
# 5bis. Excel report (one sheet per priority API + a summary sheet + a
#       shared-schema-fixes sheet) — the only report format now, and what
#       gets attached to the email.
# ============================================================================

STATUS_LABEL_FR = {
    'implemente': 'Implémenté', 'partiel': 'Partiel',
    'non_implemente': 'Non implémenté', 'a_verifier_manuellement': 'À vérifier',
}
STATUS_FILL = {
    'implemente': PatternFill('solid', fgColor='C6EFCE'),
    'partiel': PatternFill('solid', fgColor='FFEB9C'),
    'non_implemente': PatternFill('solid', fgColor='FFC7CE'),
    'a_verifier_manuellement': PatternFill('solid', fgColor='D9D9D9'),
}
HEADER_FILL = PatternFill('solid', fgColor='305496')
HEADER_FONT = Font(color='FFFFFF', bold=True)
TITLE_FONT = Font(bold=True, size=14)
WRAP = Alignment(vertical='top', wrap_text=True)


def _safe_sheet_name(name, used):
    """Excel sheet names: <=31 chars, no : \\ / ? * [ ], must be unique."""
    cleaned = re.sub(r'[:\\/?*\[\]]', '', name)[:31].strip() or 'Sheet'
    base, n = cleaned, 2
    while cleaned.lower() in used:
        suffix = f" ({n})"
        cleaned = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(cleaned.lower())
    return cleaned


def _style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_api_sheet(wb, sheet_name, path, entry):
    ws = wb.create_sheet(sheet_name)
    ws['A1'] = entry['display']
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"Endpoint : {path}"
    ws['A2'].font = Font(italic=True)

    if not entry.get('endpoint_exists_in_predig'):
        ws['A4'] = "Endpoint non documenté dans pre-dig.yaml — rien à vérifier."
        _autosize(ws, [60])
        return
    if not entry['endpoint_exists_in_data']:
        ws['A4'] = (f"Endpoint absent de data.yaml — {len(entry['fields'])} champ(s) du spec "
                     f"Mastercard non vérifiable(s).")
        ws['A4'].font = Font(bold=True, color='9C0006')

    headers = ['Champ', 'Statut', 'Fiable', 'Raisons', 'Type attendu', 'MinLength',
               'MaxLength', 'Required', 'Description', 'Origine pre-dig.yaml', 'Origine data.yaml']
    header_row = 6
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(headers))

    # reliable problems first, then minor/noisy ones, then implemented — same
    # priority ordering as the Markdown report's reliable/minor split.
    fields_sorted = sorted(
        entry['fields'],
        key=lambda f: (f['status'] == 'implemente', not f['reliable'], f['name']),
    )

    row = header_row + 1
    for f in fields_sorted:
        reason_txt = '; '.join(f['reasons'])
        values = [
            f['name'], STATUS_LABEL_FR[f['status']], 'Oui' if f['reliable'] else 'Non',
            reason_txt, f.get('type') or '', f.get('minLength') if f.get('minLength') is not None else '',
            f.get('maxLength') if f.get('maxLength') is not None else '',
            '' if f.get('required') is None else ('Oui' if f['required'] else 'Non'),
            f.get('description') or '', f.get('predig_origin') or '', f.get('data_origin') or '',
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.alignment = WRAP
            cell.fill = STATUS_FILL[f['status']]
        row += 1

    last_row = row - 1
    if last_row >= header_row + 1:
        table_ref = f"A{header_row}:{get_column_letter(len(headers))}{last_row}"
        table = Table(displayName=f"T{re.sub(r'[^A-Za-z0-9]', '', sheet_name)}", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
        ws.add_table(table)

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize(ws, [38, 16, 9, 45, 14, 11, 11, 10, 40, 22, 22])


def _write_summary_sheet(wb, audited_notes, cutoff_str, predig_direct, api_sheet_names):
    ws = wb.create_sheet("Résumé", 0)
    ws['A1'] = f"Audit historique des divergences MDES — depuis {cutoff_str}"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:D1')

    all_changes = [c for n in audited_notes for c in n['changes']]
    total = len(all_changes)
    counts = {s: sum(1 for c in all_changes if c['status'] == s) for s in STATUS_RANK}
    verifiable = total - counts['a_verifier_manuellement']
    adoption_rate = round(100 * counts['implemente'] / verifiable, 1) if verifiable else 0.0

    stats = [
        ("Changements audités (notes de release)", total),
        ("Notes couvertes", len(audited_notes)),
        ("Implémentés", counts['implemente']),
        ("Partiels", counts['partiel']),
        ("Non implémentés", counts['non_implemente']),
        ("À vérifier manuellement", counts['a_verifier_manuellement']),
        ("Taux d'adoption global (%)", adoption_rate),
    ]
    r = 3
    for label, value in stats:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Vérification directe pre-dig.yaml vs data.yaml — 7 APIs prioritaires").font = Font(bold=True, size=12)
    r += 1
    header_row = r
    headers = ['API', 'Endpoint', 'Statut', 'Écarts fiables', 'Champs total', 'Détail']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(headers))
    r += 1

    for path, display in PRIORITY_PATH_DISPLAY:
        entry = predig_direct.get(path, {})
        summary = summarize_api_status_xlsx(entry)
        ws.cell(row=r, column=1, value=display)
        ws.cell(row=r, column=2, value=path)
        status_cell = ws.cell(row=r, column=3, value=STATUS_LABEL_FR.get(summary['status'], summary['status']))
        status_cell.fill = STATUS_FILL.get(summary['status'], PatternFill())
        ws.cell(row=r, column=4, value=summary['reliable_issues'])
        ws.cell(row=r, column=5, value=summary['total_fields'])
        link_cell = ws.cell(row=r, column=6, value=f"Voir '{api_sheet_names[path]}'")
        link_cell.hyperlink = f"#'{api_sheet_names[path]}'!A1"
        link_cell.font = Font(color='0563C1', underline='single')
        r += 1

    _autosize(ws, [30, 22, 16, 14, 13, 24])


def summarize_api_status_xlsx(entry):
    """Same rollup as front/server.py's summarize_api_status() — duplicated
    here (rather than imported) since front/ imports FROM this module, not
    the other way around."""
    if not entry.get("endpoint_exists_in_predig"):
        return {"status": "unknown", "reliable_issues": 0, "total_fields": 0}
    if not entry["endpoint_exists_in_data"]:
        return {"status": "non_implemente", "reliable_issues": len(entry["fields"]), "total_fields": len(entry["fields"])}
    reliable_problems = [f for f in entry["fields"] if f["status"] != "implemente" and f["reliable"]]
    if any(f["status"] == "non_implemente" for f in reliable_problems):
        status = "non_implemente"
    elif reliable_problems:
        status = "partiel"
    else:
        status = "implemente"
    return {"status": status, "reliable_issues": len(reliable_problems), "total_fields": len(entry["fields"])}


def _write_shared_fixes_sheet(wb, predig_direct):
    field_clusters = group_shared_schema_fixes(predig_direct)
    if not field_clusters:
        return
    by_data_target = summarize_by_data_target(field_clusters)

    ws = wb.create_sheet("Écarts à cause commune")
    ws['A1'] = "Un seul schéma partagé à corriger résout plusieurs endpoints à la fois"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')

    header_row = 3
    headers = ['Schéma data.yaml', 'Champs concernés', 'Nb champs', 'Nb endpoints',
               'Endpoints concernés', 'Schémas pre-dig.yaml source']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(headers))

    r = header_row + 1
    for target in by_data_target:
        pairs_for_target = [p for p in summarize_schema_fix_clusters(field_clusters)
                             if p['data_origin'] == target['data_origin']]
        all_fields = sorted({f['name'] for p in pairs_for_target for f in p['fields']})
        ws.cell(row=r, column=1, value=target['data_origin'])
        ws.cell(row=r, column=2, value=', '.join(all_fields))
        ws.cell(row=r, column=3, value=target['field_count'])
        ws.cell(row=r, column=4, value=target['endpoint_count'])
        ws.cell(row=r, column=5, value=', '.join(target['endpoints']))
        ws.cell(row=r, column=6, value=', '.join(target['predig_origins']))
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).alignment = WRAP
        r += 1

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize(ws, [26, 40, 11, 12, 45, 30])


def _write_notes_detail_sheet(wb, audited_notes):
    ws = wb.create_sheet("Détail par note")
    headers = ['Note (version)', 'URL', 'Changement', 'Statut', 'API prioritaire',
               'Endpoint(s)', 'Champ', 'Détail statut', 'Raisons']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    r = 2
    for note in audited_notes:
        for c in sorted(note['changes'], key=lambda c: STATUS_RANK[c['status']], reverse=True):
            endpoints = ', '.join(ep['endpoint'] for ep in c['endpoint_results']) or '(endpoint non identifié)'
            field_rows = [(ep['endpoint'], f) for ep in c['endpoint_results'] for f in ep['fields']
                          if f['status'] != 'implemente']
            if not field_rows:
                values = [note['version'], note['url'], c['title'], STATUS_LABEL_FR[c['status']],
                          'Oui' if c['_is_priority'] else 'Non', endpoints, '', '', '']
                for col, v in enumerate(values, start=1):
                    ws.cell(row=r, column=col, value=v).alignment = WRAP
                r += 1
                continue
            for endpoint, f in field_rows:
                values = [note['version'], note['url'], c['title'], STATUS_LABEL_FR[c['status']],
                          'Oui' if c['_is_priority'] else 'Non', endpoint, f['name'],
                          STATUS_LABEL_FR[f['status']], '; '.join(f['reasons'])]
                for col, v in enumerate(values, start=1):
                    ws.cell(row=r, column=col, value=v).alignment = WRAP
                r += 1

    ws.freeze_panes = "A2"
    _autosize(ws, [16, 45, 40, 16, 14, 26, 30, 16, 45])


def render_report_xlsx(audited_notes, cutoff_str, xlsx_path, predig_direct=None):
    """Builds the workbook attached to the email: a 'Résumé' sheet, one sheet
    per priority API (field-by-field, colored by status), a shared-schema-
    fixes sheet, and a per-release-note detail sheet."""
    wb = Workbook()
    wb.remove(wb.active)  # placeholder default sheet, real sheets added below

    predig_direct = predig_direct or {}
    used_names = set()
    api_sheet_names = {}
    for path, display in PRIORITY_PATH_DISPLAY:
        sheet_name = _safe_sheet_name(display, used_names)
        api_sheet_names[path] = sheet_name

    _write_summary_sheet(wb, audited_notes, cutoff_str, predig_direct, api_sheet_names)

    for path, display in PRIORITY_PATH_DISPLAY:
        entry = predig_direct.get(path, {'display': display})
        _write_api_sheet(wb, api_sheet_names[path], path, entry)

    _write_shared_fixes_sheet(wb, predig_direct)
    if audited_notes:
        _write_notes_detail_sheet(wb, audited_notes)

    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    wb.save(xlsx_path)
    return xlsx_path


#  display name + canonical path per priority API, keyed by the same
#  normalized (lowercase, no slash) form used by is_priority_endpoint()
PRIORITY_DISPLAY_NAMES = {
    'requestactivationmethods': 'Request Activation Methods (RAM)',
    'deliveractivationcode': 'Deliver Activation Code (DAC)',
    'authorizeservice': 'Authorize Service (AS)',
    'notifyserviceactivated': 'Notify Service Activated (NSA)',
    'notifytokenupdated': 'Notify Token Updated (NTU)',
    'validateactivationcode': 'Validate Activation Code (VAC)',
    'getaccountinformation': 'Get Account Information (GAI)',
}
PRIORITY_DISPLAY_ORDER = [
    'requestactivationmethods', 'deliveractivationcode', 'authorizeservice',
    'notifyserviceactivated', 'notifytokenupdated',
    'validateactivationcode', 'getaccountinformation',
]


def _endpoint_norm_key(path):
    return re.sub(r'[^a-z]', '', path.lower())


def summarize_priority_api(entry):
    """One-line summary for the email body ('Request Activation Methods (RAM) :
    response — 2 champ(s) manquant(s)') — split request/response, counting only
    reliable (missing/type/maxLength) mismatches. Field-by-field detail is left
    to the attached report, not repeated here. Returns None when the API has
    nothing to report (fully implemented, or not documented in pre-dig.yaml)."""
    if not entry.get('endpoint_exists_in_predig'):
        return None
    if not entry['endpoint_exists_in_data']:
        return f"endpoint absent de data.yaml ({len(entry['fields'])} champ(s) non vérifiable(s))"
    reliable_problems = [f for f in entry['fields'] if f['status'] != 'implemente' and f['reliable']]
    if not reliable_problems:
        return None
    counts = {}
    for f in reliable_problems:
        side = 'response' if f['name'].startswith('response') else 'request'
        bucket = counts.setdefault(side, {'manquant': 0, 'divergent': 0})
        bucket['manquant' if f['status'] == 'non_implemente' else 'divergent'] += 1
    parts = []
    for side in ('request', 'response'):
        if side not in counts:
            continue
        sub = []
        if counts[side]['manquant']:
            sub.append(f"{counts[side]['manquant']} champ(s) manquant(s)")
        if counts[side]['divergent']:
            sub.append(f"{counts[side]['divergent']} champ(s) divergent(s)")
        parts.append(f"{side} — {', '.join(sub)}")
    return " ; ".join(parts)


def render_email_draft(audited_notes, cutoff_str, predig_direct=None):
    all_changes = [c for n in audited_notes for c in n['changes']]
    urgent = [c for c in all_changes if c['status'] != 'implemente']

    # Group every non-implemented/partial FIELD of every priority-API change
    # by the actual endpoint it belongs to, so the email names the API and
    # its fields directly instead of just listing change titles.
    by_endpoint = {}  # norm_key -> [(field_name, field_status, note_version, change_title), ...]
    for c in urgent:
        for ep in c['endpoint_results']:
            key = _endpoint_norm_key(ep['endpoint'])
            if key not in PRIORITY_DISPLAY_NAMES:
                continue
            if not ep['endpoint_exists']:
                by_endpoint.setdefault(key, []).append(
                    (None, 'endpoint absent de data.yaml', c['note_version'], c['title']))
                continue
            for f in ep['fields']:
                if f['status'] == 'implemente':
                    continue
                by_endpoint.setdefault(key, []).append(
                    (f['name'], f['status'], c['note_version'], c['title']))

    lines = [
        f"Objet : [MDES] Audit divergences — {len(urgent)} écart(s) à traiter (depuis {cutoff_str})",
        f"À : {EMAIL_RECIPIENTS or '(destinataires non définis — placeholder, voir MC_DIVERGENCE_RECIPIENTS)'}",
        "",
        f"Bonjour,", "",
        f"L'audit historique des pre-release notes Mastercard MDES (depuis {cutoff_str}) a détecté "
        f"{len(urgent)} changement(s) partiellement ou non implémentés dans data.yaml sur "
        f"{len(all_changes)} changements au total.",
        "",
        "Le rapport complet (détail note par note) est joint en pièce jointe à cet email.",
        "",
    ]

    if predig_direct:
        lines.append("### APIs prioritaires — résumé (vérification directe pre-dig.yaml vs data.yaml, "
                     "écarts fiables uniquement — détail champ par champ en pièce jointe)")
        lines.append("")
        any_summary = False
        for path, entry in predig_direct.items():
            summary = summarize_priority_api(entry)
            if summary is None:
                continue
            any_summary = True
            lines.append(f"- **{entry['display']}** : {summary}")
        if not any_summary:
            lines.append("Aucun écart fiable détecté sur les APIs prioritaires.")
        lines.append("")
        lines.append("(Divergences mineures minLength/required omises du résumé, probablement "
                     "dues au style de génération de data.yaml — voir le rapport complet en pièce jointe "
                     "pour le détail.)")
        lines.append("")

        field_clusters = group_shared_schema_fixes(predig_direct)
        by_data_target = summarize_by_data_target(field_clusters)
        if by_data_target:
            lines.append("### Écarts à cause commune — corriger une fois pour résoudre plusieurs endpoints "
                         "(détail en pièce jointe)")
            lines.append("")
            for r in by_data_target:
                lines.append(f"- Corriger `{r['data_origin']}` dans data.yaml → résout {r['field_count']} "
                             f"champ(s) sur {r['endpoint_count']} endpoint(s) d'un coup : "
                             f"{', '.join(r['endpoints'])}")
            lines.append("")

    if by_endpoint:
        lines.append("### Contexte par note (quelle release a introduit quoi — indicatif, détail en pièce jointe)")
        lines.append("")
        for key in PRIORITY_DISPLAY_ORDER:
            entries = by_endpoint.get(key)
            if not entries:
                continue
            notes = sorted({f"{note_version} ({title})" for _, _, note_version, title in entries})
            lines.append(f"- **{PRIORITY_DISPLAY_NAMES[key]}** : {len(entries)} champ(s) concerné(s) "
                         f"sur {len(notes)} note(s) — {'; '.join(notes)}")
        lines.append("")

    to_verify = [c for c in urgent if c['status'] == 'a_verifier_manuellement']
    if to_verify:
        lines.append("### À vérifier manuellement (note sans tableau structuré — rien de vérifiable automatiquement)")
        lines.append("")
        for c in to_verify:
            endpoints = ', '.join(ep['endpoint'] for ep in c['endpoint_results']) or '(endpoint non identifié)'
            lines.append(f"- {c['note_version']} — **{c['title']}** ({endpoints})")
        lines.append("")

    other_urgent = [c for c in urgent if not c['_is_priority'] and c['status'] != 'a_verifier_manuellement']
    if other_urgent:
        lines.append("### Autres écarts (APIs non prioritaires)")
        lines.append("")
        for c in other_urgent:
            lines.append(f"- [{c['status']}] {c['note_version']} — {c['title']}")
        lines.append("")

    return '\n'.join(lines)

# ============================================================================
# 6. Orchestration
# ============================================================================

def is_priority_endpoint(path):
    return re.sub(r'[^a-z]', '', path.lower()) in {
        p.replace('/', '') for p in PRIORITY_PATHS
    }


def run(data_yaml_path, cutoff, cache_dir, xlsx_path, refresh, send=False):
    """Returns (audited_notes, subject, body). No .md file is written
    anywhere in this pipeline anymore — the Excel workbook (xlsx_path) is
    the only report artifact, and the email subject/body are handed back
    directly instead of being round-tripped through a local draft file."""
    os.makedirs(cache_dir, exist_ok=True)
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

    if refresh or not os.path.exists(os.path.join(cache_dir, 'known_releases.json')):
        fetch_mc_source.fetch_release_notes(cache_dir)
    # pre-dig.yaml is ALWAYS re-fetched fresh from Mastercard, never just
    # reused from whatever was cached last time — this is the authoritative
    # spec the direct diff runs against, so a new Mastercard release must be
    # picked up on the very next run with no manual --refresh step needed.
    predig_path = os.path.join(cache_dir, 'pre-dig.yaml')
    fetch_mc_source.fetch_pre_dig_spec(cache_dir)

    releases = load_releases_since(cache_dir, cutoff)
    print(f"  [phase1] {len(releases)} release(s) depuis {cutoff}", file=sys.stderr)

    print(f"  [phase1] loading data.yaml <- {data_yaml_path}", file=sys.stderr)
    data_spec = load_spec(data_yaml_path, repair=True)
    endpoint_cache = {}

    print(f"  [phase1] loading pre-dig.yaml <- {predig_path}", file=sys.stderr)
    predig_spec = load_spec(predig_path, repair=True)
    print("  [phase1] direct structural diff: pre-dig.yaml vs data.yaml (7 priority APIs)", file=sys.stderr)
    predig_direct = audit_predig_vs_data_direct(predig_spec, data_spec)

    audited_notes = []
    for entry in releases:
        text, _ = get_note_markdown(entry['url'], cache_dir)
        record = parse_note(text, url=entry['url'])
        record['mdes_release'] = entry['mdes_release']
        record['published_date'] = entry['upgrade_date']

        audited_changes = []
        for c in record['changes']:
            endpoint_results = audit_change(c, data_spec, endpoint_cache)
            status = worst_status([r['status'] for r in endpoint_results])
            is_priority = any(is_priority_endpoint(r['endpoint']) for r in endpoint_results)
            audited_changes.append({
                **c,
                'endpoint_results': endpoint_results,
                'status': status,
                '_is_priority': is_priority,
                'note_version': record['version'],
            })

        # priority-first ordering within the note too (prompt step 3)
        audited_changes.sort(key=lambda c: (not c['_is_priority'], -STATUS_RANK[c['status']]))

        audited_notes.append({
            'version': record['version'],
            'url': record['url'],
            'mdes_release': record['mdes_release'],
            'published_date': record['published_date'],
            'timeline': record['timeline'],
            'changes': audited_changes,
        })

    render_report_xlsx(audited_notes, f"{cutoff[0]}-{cutoff[1]:02d}", xlsx_path, predig_direct)
    print(f"  [phase1] report (xlsx) -> {xlsx_path}", file=sys.stderr)

    email = render_email_draft(audited_notes, f"{cutoff[0]}-{cutoff[1]:02d}", predig_direct)
    subject_line, _, body = email.partition('\n')
    subject = subject_line.removeprefix('Objet : ').strip()
    body = body.lstrip('\n')

    if send:
        try:
            sent_to = send_email_module.send_email(subject, body, attachments=[xlsx_path])
            print(f"  [phase1] email SENT to {', '.join(sent_to)} with {os.path.basename(xlsx_path)} attached",
                  file=sys.stderr)
        except send_email_module.SmtpConfigError as e:
            print(f"  [phase1] email NOT sent — {e}", file=sys.stderr)
            raise
    else:
        print(f"  [phase1] xlsx report -> {xlsx_path} (NOT sent — pass --send to actually email it)",
              file=sys.stderr)

    return audited_notes, subject, body


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('--data-yaml', default=DEFAULT_DATA_YAML)
    parser.add_argument('--cutoff', default='2025-01', help="YYYY-MM, default 2025-01")
    parser.add_argument('--cache-dir', default=fetch_mc_source.DEFAULT_CACHE_DIR)
    parser.add_argument('--report', default=DEFAULT_REPORT_XLSX_PATH, help="Excel report output path")
    parser.add_argument('--refresh', action='store_true',
                         help="Re-fetch known_releases.json even if already cached "
                              "(pre-dig.yaml is always re-fetched fresh regardless of this flag)")
    parser.add_argument('--send', action='store_true',
                         help="Actually email the report via SMTP (needs SMTP_HOST/SMTP_USER/"
                              "SMTP_PASSWORD env vars — fails loudly if missing). Without this "
                              "flag, only the local .xlsx report is written.")
    args = parser.parse_args()

    cutoff = parse_cutoff(args.cutoff)
    try:
        run(args.data_yaml, cutoff, args.cache_dir, args.report, args.refresh, send=args.send)
    except send_email_module.SmtpConfigError as e:
        print(f"ERREUR : {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
