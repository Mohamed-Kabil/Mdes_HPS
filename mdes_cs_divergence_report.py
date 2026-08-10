#!/usr/bin/env python3
"""
mdes_cs_divergence_report.py — compares the official MDES Customer Service
OpenAPI spec (Mastercard's published truth, fetched from
developer.mastercard.com) against what the pwc-api-sp5_api Java codebase
actually builds (extract_java_field_mapping.py's generated output), for the
6 CS operations (Search, Token Activate, Update, Suspend, Unsuspend, Delete).

Same relationship as pre-dig.yaml vs data.yaml in mc_divergence/ — the
official spec plays pre-dig.yaml's role (source of truth for what the API
supports), the Java extraction plays data.yaml's role (what's actually
implemented) — just applied to the CS API domain instead of
Pre-Digitization, and sourced from Java instead of a second spec.

Field matching: exact dotted-path match first (the Java DTOs were built
against this same spec family, so names usually agree exactly), falling
back to last-segment name match tolerant of case and known spelling
variants (see SPELLING_FOLDS — same idea as check_field_changes.py's
cypher/cipher fold).

Per-field status:
  non_implemente  — no Java field found anywhere in any variant
  non_verifiable  — matched, but the Java extraction itself couldn't
                    resolve that field (source: unresolved) — a gap in the
                    input data, not a confirmed absence (see
                    MDES_CS_API_JAVA_MAPPING_LINK.md)
  partiel         — matched via the local LLM's inference rather than a
                    direct mechanical scan (source: llm-inferred)
  implemente      — matched directly by the mechanical scanner

Usage:
    python3 mdes_cs_divergence_report.py
        [--spec C:\\Users\\moham\\Desktop\\input\\mdes-customer-service.yaml]
        [--java-mapping <path to mdes_cs_api_schemas.generated.json>]
        [-o mdes_cs_divergence_report.json]
"""

import argparse
import json
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(HERE, 'mc_divergence'))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from diff_openapi_all import load_spec, get_operation, flatten_schema
from diff_predig_vs_data import extract_request_schema
from phase1_historical_audit import (
    _safe_sheet_name, _style_header_row, _autosize,
    HEADER_FILL, HEADER_FONT, TITLE_FONT, WRAP,
)
import send_email as send_email_module

DEFAULT_CS_SPEC_YAML = os.environ.get(
    'MDES_CS_SPEC_YAML', r'C:\Users\moham\Desktop\input\mdes-customer-service.yaml')
DEFAULT_CS_JAVA_MAPPING_JSON = os.environ.get(
    'MDES_CS_MAPPING_JSON',
    r'C:\Users\moham\Downloads\pwc-api-sp5_api\Mdes_cs_api\generated\mdes_cs_api_schemas.generated.json')
DEFAULT_OUTPUT = os.path.join(HERE, 'mdes_cs_divergence_report.json')
DEFAULT_OUTPUT_MD = os.path.join(HERE, 'mdes_cs_divergence_report.md')
DEFAULT_REPORT_XLSX_PATH = os.path.join(HERE, 'mc_divergence', 'reports', 'mdes_cs_divergence_report.xlsx')

STATUS_LABEL_FR = {
    'implemente': 'Implémenté', 'partiel': 'Partiel',
    'non_verifiable': 'Non vérifiable', 'non_implemente': 'Non implémenté',
}
STATUS_FILL = {
    'implemente': PatternFill('solid', fgColor='C6EFCE'),
    'partiel': PatternFill('solid', fgColor='FFEB9C'),
    'non_verifiable': PatternFill('solid', fgColor='D9D9D9'),
    'non_implemente': PatternFill('solid', fgColor='FFC7CE'),
}

OPERATIONS = [
    ('Search', '/{id}/search'),
    ('Token Activate', '/{id}/token/activate'),
    ('Token Update', '/{id}/token/update'),
    ('Token Suspend', '/{id}/token/suspend'),
    ('Token Unsuspend', '/{id}/token/unsuspend'),
    ('Token Delete', '/{id}/token/delete'),
]

SPELLING_FOLDS = {
    'organisation': 'organization',
}


def normalize_leaf(name):
    n = re.sub(r'[^a-z0-9]', '', name.lower())
    return SPELLING_FOLDS.get(n, n)


def strip_wrapper_prefix(fields):
    """flatten_schema() includes the single top-level wrapper key (e.g.
    'SearchRequest') as its own '.'-free entry, with every real field
    nested under 'SearchRequest.'; the Java-side field paths don't carry
    this wrapper. Strip it so both sides line up."""
    tops = [k for k in fields if '.' not in k]
    if len(tops) != 1:
        return dict(fields)
    prefix = tops[0] + '.'
    return {k[len(prefix):]: v for k, v in fields.items() if k.startswith(prefix)}


def flatten_java_fields(fields, prefix=''):
    """The generated mapping nests object-typed fields under their own
    'properties' dict (e.g. AuditInfo -> properties -> UserId/UserName/...)
    rather than using dotted keys directly. Flatten to the same
    'Parent.Child' convention flatten_schema() uses on the spec side, so
    both sides compare on equal footing."""
    out = {}
    for name, info in fields.items():
        full = f'{prefix}{name}'
        out[full] = {
            'type': info.get('type'),
            'source': info.get('source'),
            'javaSource': info.get('javaSource'),
            'javaExpr': info.get('javaExpr'),
        }
        if info.get('properties'):
            out.update(flatten_java_fields(info['properties'], prefix=full + '.'))
    return out


def load_java_fields(mapping_json_path):
    """{operation_name: {field_path: [{'variant', 'source', 'javaSource'}]}}
    — a field can appear in multiple variants (e.g. Search's 3 request
    shapes), each independently tagged mechanical-scan/llm-inferred/unresolved."""
    with open(mapping_json_path, encoding='utf-8') as f:
        raw = json.load(f)
    out = {}
    for op in raw.get('operations', []):
        by_field = {}
        for variant_name, variant in op.get('variants', {}).items():
            flat = flatten_java_fields(variant.get('fields', {}))
            for field_path, info in flat.items():
                by_field.setdefault(field_path, []).append({
                    'variant': variant_name,
                    'source': info.get('source'),
                    'javaSource': info.get('javaSource'),
                })
        out[op['operation']] = by_field
    return out, raw.get('_meta', {})


def match_field(spec_path, java_fields_by_path):
    if spec_path in java_fields_by_path:
        return spec_path, java_fields_by_path[spec_path]
    spec_leaf = normalize_leaf(spec_path.split('.')[-1])
    for jpath, hits in java_fields_by_path.items():
        if normalize_leaf(jpath.split('.')[-1]) == spec_leaf:
            return jpath, hits
    return None, None


def audit_operation(spec_fields, java_fields_by_path):
    results = []
    for spec_path, info in sorted(spec_fields.items()):
        matched_path, hits = match_field(spec_path, java_fields_by_path)
        if hits is None:
            status = 'non_implemente'
        elif any(h['source'] == 'unresolved' for h in hits):
            status = 'non_verifiable'
        elif any(h['source'] == 'llm-inferred' for h in hits):
            status = 'partiel'
        else:
            status = 'implemente'
        results.append({
            'field': spec_path,
            'status': status,
            'required': info.get('required'),
            'type': info.get('type'),
            'matched_java_field': matched_path,
            'variants': sorted({h['variant'] for h in hits}) if hits else [],
            'java_source': sorted({h['javaSource'] for h in hits if h.get('javaSource')}) if hits else [],
        })
    return results


def group_shared_field_gaps(operations):
    """Fields marked non_implemente identically across 2+ operations at
    once — same relationship as phase1_historical_audit.py's
    group_shared_schema_fixes(), but clustering directly on the field path
    (there's no separate spec-schema-origin vs data-schema-origin pair here
    like the pre-dig/data.yaml side has — the CS spec's own field path IS
    the shared key, since the exact same dotted path showing up as missing
    on 2+ operations means the exact same underlying Java gap)."""
    by_field = {}
    for op in operations:
        if 'error' in op:
            continue
        for f in op.get('fields', []):
            if f['status'] != 'non_implemente':
                continue
            entry = by_field.setdefault(f['field'], {
                'field': f['field'], 'required': f['required'], 'type': f['type'], 'operations': [],
            })
            entry['operations'].append(op['operation'])
    clusters = [c for c in by_field.values() if len(c['operations']) >= 2]
    clusters.sort(key=lambda c: len(c['operations']), reverse=True)
    return clusters


def summarize_by_root_field(field_clusters):
    """Rolls per-field clusters up to their top-level parent (e.g. every
    'EncryptedAccountInformation.*' leaf rolls up to 'EncryptedAccountInformation')
    — the actionable unit: one Java change (wiring up that object) resolves
    every leaf under it, across every operation missing it. Same idea as
    summarize_by_data_target() on the pre-dig side."""
    rollup = {}
    for c in field_clusters:
        root = c['field'].split('.')[0]
        r = rollup.setdefault(root, {'root_field': root, 'fields': [], 'operations': set()})
        r['fields'].append(c['field'])
        r['operations'].update(c['operations'])
    out = []
    for r in rollup.values():
        out.append({
            'root_field': r['root_field'],
            'field_count': len(r['fields']),
            'fields': sorted(r['fields']),
            'operation_count': len(r['operations']),
            'operations': sorted(r['operations']),
        })
    out.sort(key=lambda r: (r['operation_count'], r['field_count']), reverse=True)
    return out


def run(spec_path, java_mapping_path):
    spec = load_spec(spec_path, repair=True)
    java_fields_by_op, java_meta = load_java_fields(java_mapping_path)

    operations = []
    for name, path in OPERATIONS:
        op = get_operation(spec, path, 'post')
        if op is None:
            operations.append({'operation': name, 'path': path, 'error': 'path absent du spec officiel'})
            continue
        schema, _ = extract_request_schema(op, spec)
        spec_fields = strip_wrapper_prefix(flatten_schema(schema))
        java_fields_by_path = java_fields_by_op.get(name, {})
        fields = audit_operation(spec_fields, java_fields_by_path)

        counts = {}
        for f in fields:
            counts[f['status']] = counts.get(f['status'], 0) + 1

        operations.append({
            'operation': name,
            'path': path,
            'total_fields': len(fields),
            'counts': counts,
            'fields': fields,
        })

    shared_gaps = summarize_by_root_field(group_shared_field_gaps(operations))

    return {
        'spec_source': spec_path,
        'java_mapping_source': java_mapping_path,
        'java_mapping_generated_at': java_meta.get('generatedAt'),
        'operations': operations,
        'shared_gaps': shared_gaps,
    }


def render_markdown(report):
    lines = [
        '# MDES Customer Service — écarts spec officiel vs implémentation Java (auto-généré)',
        '',
        f"Spec officiel : `{report['spec_source']}`",
        f"Extraction Java : `{report['java_mapping_source']}` (générée le {report['java_mapping_generated_at']})",
        '',
        '`non_implemente` = champ du spec absent de tout ce que le code Java construit actuellement. '
        '`non_verifiable` = un champ Java correspondant existe mais son identité n\'a pas pu être résolue '
        '(voir MDES_CS_API_JAVA_MAPPING_LINK.md) — écart dans la donnée, pas une absence confirmée. '
        '`partiel` = résolu par le modèle local plutôt que par un scan direct. `implemente` = trouvé directement.',
        '',
    ]

    shared_gaps = report.get('shared_gaps') or []
    if shared_gaps:
        lines.append('## Écarts à cause commune — corriger une fois pour résoudre plusieurs opérations')
        lines.append('')
        for g in shared_gaps:
            lines.append(
                f"- Corriger `{g['root_field']}` côté Java → résout **{g['field_count']} champ(s)** "
                f"sur **{g['operation_count']} opération(s)** d'un coup : {', '.join(g['operations'])}"
            )
            for field in g['fields']:
                lines.append(f"  - `{field}`")
        lines.append('')

    for op in report['operations']:
        if 'error' in op:
            lines.append(f"## {op['operation']} — {op['path']}\n\n{op['error']}\n")
            continue
        c = op['counts']
        lines.append(f"## {op['operation']} — `POST {op['path']}`")
        lines.append(
            f"{op['total_fields']} champ(s) au total — "
            f"{c.get('implemente', 0)} implémenté(s), {c.get('partiel', 0)} partiel(s), "
            f"{c.get('non_verifiable', 0)} non vérifiable(s), {c.get('non_implemente', 0)} non implémenté(s)."
        )
        problems = [f for f in op['fields'] if f['status'] != 'implemente']
        if problems:
            lines.append('')
            lines.append('| Champ (spec officiel) | Statut | Requis | Champ Java correspondant |')
            lines.append('|---|---|---|---|')
            for f in problems:
                lines.append(
                    f"| `{f['field']}` | {f['status']} | {'oui' if f['required'] else 'non'} | "
                    f"{f['matched_java_field'] or '—'} |"
                )
        lines.append('')
    return '\n'.join(lines)


# ============================================================================
# Excel report — same styling/helpers as phase1_historical_audit.py's report,
# reused directly rather than reimplemented (_safe_sheet_name, _style_header_row,
# _autosize, HEADER_FILL/HEADER_FONT/TITLE_FONT/WRAP).
# ============================================================================

def _write_summary_sheet(wb, report, op_sheet_names):
    ws = wb.create_sheet("Résumé", 0)
    ws['A1'] = "MDES Customer Service — écarts spec officiel vs implémentation Java"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')
    ws['A2'] = f"Spec officiel : {report['spec_source']}"
    ws['A3'] = f"Extraction Java du {report['java_mapping_generated_at']}"

    headers = ['Opération', 'Total champs', 'Implémenté', 'Partiel', 'Non vérifiable', 'Non implémenté']
    header_row = 5
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(headers))

    row = header_row + 1
    for op in report['operations']:
        if 'error' in op:
            continue
        c = op['counts']
        values = [op['operation'], op['total_fields'], c.get('implemente', 0), c.get('partiel', 0),
                  c.get('non_verifiable', 0), c.get('non_implemente', 0)]
        for col, v in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=v)
        sheet_name = op_sheet_names.get(op['operation'])
        if sheet_name:
            ws.cell(row=row, column=1).hyperlink = f"#'{sheet_name}'!A1"
            ws.cell(row=row, column=1).font = Font(underline='single', color='0563C1')
        row += 1

    _autosize(ws, [22, 14, 14, 12, 16, 16])


def _write_operation_sheet(wb, sheet_name, op):
    ws = wb.create_sheet(sheet_name)
    ws['A1'] = op['operation']
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"POST {op['path']}"
    ws['A2'].font = Font(italic=True)

    headers = ['Champ (spec officiel)', 'Statut', 'Requis', 'Type', 'Champ Java correspondant',
               'Variante(s)', 'Source Java']
    header_row = 4
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(headers))

    fields_sorted = sorted(op['fields'], key=lambda f: (f['status'] == 'implemente', f['field']))
    row = header_row + 1
    for f in fields_sorted:
        values = [
            f['field'], STATUS_LABEL_FR[f['status']], 'Oui' if f['required'] else 'Non',
            f.get('type') or '', f.get('matched_java_field') or '—',
            ', '.join(f.get('variants') or []), '; '.join(f.get('java_source') or []),
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.alignment = WRAP
            cell.fill = STATUS_FILL[f['status']]
        row += 1

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize(ws, [45, 16, 9, 12, 45, 30, 60])


def _write_shared_gaps_sheet(wb, shared_gaps):
    if not shared_gaps:
        return
    ws = wb.create_sheet("Écarts à cause commune")
    ws['A1'] = "Un seul changement côté Java résout plusieurs opérations à la fois"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:D1')

    headers = ['Champ racine', 'Champs concernés', 'Nb champs', 'Nb opérations', 'Opérations concernées']
    header_row = 3
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(headers))

    row = header_row + 1
    for g in shared_gaps:
        values = [g['root_field'], ', '.join(g['fields']), g['field_count'], g['operation_count'],
                  ', '.join(g['operations'])]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.alignment = WRAP
        row += 1

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize(ws, [26, 60, 11, 13, 40])


def render_report_xlsx(report, xlsx_path):
    wb = Workbook()
    wb.remove(wb.active)

    used_names = set()
    op_sheet_names = {}
    for op in report['operations']:
        if 'error' in op:
            continue
        sheet_name = _safe_sheet_name(op['operation'], used_names)
        op_sheet_names[op['operation']] = sheet_name

    _write_summary_sheet(wb, report, op_sheet_names)
    for op in report['operations']:
        if 'error' in op:
            continue
        _write_operation_sheet(wb, op_sheet_names[op['operation']], op)
    _write_shared_gaps_sheet(wb, report.get('shared_gaps') or [])

    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    wb.save(xlsx_path)


# ============================================================================
# Email draft — same relationship to render_report_xlsx() as
# phase1_historical_audit.render_email_draft() has to its own xlsx report:
# short readable summary, full detail travels as the attached workbook.
# ============================================================================

def email_subject(report):
    total_issues = sum(
        (op['counts'].get('non_implemente', 0) + op['counts'].get('non_verifiable', 0)
         + op['counts'].get('partiel', 0))
        for op in report['operations'] if 'error' not in op
    )
    return f"[MDES Customer Service] Audit divergences — {total_issues} écart(s) à traiter"


def render_email_body(report):
    lines = [
        "Bonjour,", "",
        "L'audit MDES Customer Service (spec officiel Mastercard vs implémentation Java) "
        "a été relancé. Résumé ci-dessous, détail champ par champ en pièce jointe.",
        "",
    ]

    for op in report['operations']:
        if 'error' in op:
            lines.append(f"- **{op['operation']}** : {op['error']}")
            continue
        c = op['counts']
        issues = c.get('non_implemente', 0) + c.get('non_verifiable', 0) + c.get('partiel', 0)
        lines.append(f"- **{op['operation']}** : {issues} écart(s) sur {op['total_fields']} champ(s) "
                     f"({c.get('non_implemente', 0)} non implémenté(s), "
                     f"{c.get('non_verifiable', 0)} non vérifiable(s), {c.get('partiel', 0)} partiel(s))")
    lines.append('')

    shared_gaps = report.get('shared_gaps') or []
    if shared_gaps:
        lines.append("### Écarts à cause commune — corriger une fois pour résoudre plusieurs opérations")
        lines.append('')
        for g in shared_gaps:
            lines.append(f"- Corriger `{g['root_field']}` côté Java → résout {g['field_count']} champ(s) "
                         f"sur {g['operation_count']} opération(s) d'un coup : {', '.join(g['operations'])}")
        lines.append('')

    return '\n'.join(lines)


def main():
    parser = argparse.ArgumentParser(
        description='Compare le spec officiel MDES Customer Service a ce que le code Java construit.'
    )
    parser.add_argument('--spec', default=DEFAULT_CS_SPEC_YAML)
    parser.add_argument('--java-mapping', default=DEFAULT_CS_JAVA_MAPPING_JSON)
    parser.add_argument('-o', '--output', default=DEFAULT_OUTPUT)
    parser.add_argument('--output-md', default=DEFAULT_OUTPUT_MD)
    args = parser.parse_args()

    if not os.path.exists(args.spec):
        print(f"ERROR: spec introuvable: {args.spec}", file=sys.stderr)
        sys.exit(1)
    if not os.path.exists(args.java_mapping):
        print(f"ERROR: mapping Java introuvable: {args.java_mapping}", file=sys.stderr)
        sys.exit(1)

    report = run(args.spec, args.java_mapping)

    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    with open(args.output_md, 'w', encoding='utf-8') as f:
        f.write(render_markdown(report))

    for op in report['operations']:
        if 'error' in op:
            print(f"  [FAIL] {op['operation']}: {op['error']}", file=sys.stderr)
            continue
        c = op['counts']
        print(f"  [ok] {op['operation']}: {op['total_fields']} champ(s) — "
              f"{c.get('non_implemente', 0)} non_implemente, {c.get('non_verifiable', 0)} non_verifiable, "
              f"{c.get('partiel', 0)} partiel", file=sys.stderr)

    print(f"[report] {args.output}", file=sys.stderr)
    print(f"[report] {args.output_md}", file=sys.stderr)


if __name__ == '__main__':
    main()
